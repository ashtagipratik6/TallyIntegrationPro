import sys
import re
import html
import time
import calendar
from datetime import datetime
import requests
from PySide6.QtCore import (
    Qt, QThread, Signal, QPoint, QRect, QSize, QTimer, QDate
)
from PySide6.QtGui import (
    QFont, QColor, QBrush, QPen, QPainter, QPolygon
)
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QComboBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QStyledItemDelegate,
    QCompleter, QCheckBox, QFrame, QTextEdit, QMessageBox,
    QDialog, QListWidget, QListWidgetItem, QMenu, QDateEdit,
    QFileDialog, QFormLayout, QGroupBox, QTabWidget, QDialogButtonBox,
    QStatusBar, QProgressBar, QSpinBox, QDoubleSpinBox, QCalendarWidget
)

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    openpyxl = None
    Workbook = None

DEFAULT_TALLY_URL = "http://localhost:9000"
CURRENT_VERSION = "v1.0.0"

# ============================================================================
#  0. AUTO-UPDATER WORKERS
# ============================================================================
class UpdateCheckerWorker(QThread):
    update_available = Signal(str, str) # version, download_url

    def run(self):
        try:
            url = "https://api.github.com/repos/ashtagipratik6/TallyIntegrationPro/releases/latest"
            resp = requests.get(url, timeout=5)
            if resp.status_code == 200:
                data = resp.json()
                latest_version = data.get("tag_name", "")
                if latest_version and latest_version != CURRENT_VERSION:
                    assets = data.get("assets", [])
                    if assets:
                        download_url = assets[0]["browser_download_url"]
                        self.update_available.emit(latest_version, download_url)
        except Exception:
            pass

class UpdateDownloaderWorker(QThread):
    progress = Signal(int)
    finished_download = Signal(str)
    error = Signal(str)

    def __init__(self, download_url):
        super().__init__()
        self.download_url = download_url

    def run(self):
        try:
            import os
            resp = requests.get(self.download_url, stream=True, timeout=10)
            resp.raise_for_status()
            total_size = int(resp.headers.get('content-length', 0))
            
            temp_path = os.path.join(os.environ.get('TEMP', ''), "TallyIntegrationPro_Update.exe")
            downloaded = 0
            with open(temp_path, 'wb') as f:
                for chunk in resp.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        if total_size:
                            self.progress.emit(int((downloaded / total_size) * 100))
            self.finished_download.emit(temp_path)
        except Exception as e:
            self.error.emit(str(e))

# ============================================================================
#  1. DATA HELPERS & XML FORMATTERS (Original + Create)
# ============================================================================
def clean_float(val):
    if not val:
        return 0.0
    cleaned = re.sub(r"[^\d.-]", "", str(val))
    try:
        return abs(float(cleaned))
    except ValueError:
        return 0.0

def format_date_to_ui(date_str):
    d = str(date_str).strip()
    if len(d) == 8 and d.isdigit():
        return f"{d[6:8]}-{d[4:6]}-{d[0:4]}"
    return date_str

def format_date_to_tally(date_str):
    d = str(date_str).strip()
    parts = re.split(r"[-/]", d)
    if len(parts) == 3 and len(parts[2]) == 4:
        return f"{parts[2]}{parts[1].zfill(2)}{parts[0].zfill(2)}"
    return d.replace("-", "").replace("/", "")

def normalize_tally_date(date_str):
    d = str(date_str).strip()
    if re.fullmatch(r"\d{8}", d):
        return d
    for fmt in ("%d-%b-%Y", "%d-%b-%y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(d, fmt).strftime("%Y%m%d")
        except ValueError:
            pass
    return d

def escape_xml(text):
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )

def strip_invalid_xml_chars(text):
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", str(text))

def generate_month_ranges(start_str, end_str):
    dt_start = datetime.strptime(start_str, "%Y%m%d")
    dt_end = datetime.strptime(end_str, "%Y%m%d")
    ranges = []
    curr = dt_start
    while curr <= dt_end:
        _, last_day = calendar.monthrange(curr.year, curr.month)
        m_end = datetime(curr.year, curr.month, last_day)
        if m_end > dt_end:
            m_end = dt_end
        ranges.append((curr.strftime("%Y%m%d"), m_end.strftime("%Y%m%d")))
        if curr.month == 12:
            curr = datetime(curr.year + 1, 1, 1)
        else:
            curr = datetime(curr.year, curr.month + 1, 1)
    return ranges

# ============================================================================
#  2. TALLY XML PROTOCOL BUILDERS (Original + Create)
# ============================================================================
def build_active_company_request():
    return """<ENVELOPE>
  <HEADER>
    <VERSION>1</VERSION>
    <TALLYREQUEST>Export</TALLYREQUEST>
    <TYPE>Collection</TYPE>
    <ID>ActiveCompanyAndPeriod</ID>
  </HEADER>
  <BODY>
    <DESC>
      <STATICVARIABLES>
        <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
      </STATICVARIABLES>
      <TDL>
        <TDLMESSAGE>
          <COLLECTION NAME="ActiveCompanyAndPeriod">
            <TYPE>Company</TYPE>
            <FETCH>Name</FETCH>
            <COMPUTE>PeriodFrom : ##SVFROMDATE</COMPUTE>
            <COMPUTE>PeriodTo : ##SVTODATE</COMPUTE>
            <FILTER>IsActiveCompanyFilter</FILTER>
          </COLLECTION>
          <SYSTEM TYPE="Formula" NAME="IsActiveCompanyFilter">
            $Name = ##SVCURRENTCOMPANY
          </SYSTEM>
        </TDLMESSAGE>
      </TDL>
    </DESC>
  </BODY>
</ENVELOPE>"""

def build_ledger_list_request(company=""):
    company_var = f"<SVCURRENTCOMPANY>{escape_xml(company)}</SVCURRENTCOMPANY>" if company else ""
    return f"""<ENVELOPE>
  <HEADER>
    <VERSION>1</VERSION>
    <TALLYREQUEST>EXPORT</TALLYREQUEST>
    <TYPE>COLLECTION</TYPE>
    <ID>ListOfLedgers</ID>
  </HEADER>
  <BODY>
    <DESC>
      <STATICVARIABLES>
        <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
        {company_var}
      </STATICVARIABLES>
    </DESC>
  </BODY>
</ENVELOPE>"""

def extract_ledger_names_from_xml(raw_xml):
    raw_names = re.findall(r"<NAME\b[^>]*>(.*?)</NAME>", raw_xml, flags=re.DOTALL | re.IGNORECASE)
    if not raw_names:
        raw_names = re.findall(r'NAME="([^"]+)"', raw_xml)
    cleaned_set = set()
    for n in raw_names:
        clean_str = re.sub(r"<[^>]+>", "", n)
        clean_str = html.unescape(clean_str.strip())
        if clean_str and not clean_str.startswith("$$"):
            cleaned_set.add(clean_str)
    return sorted(list(cleaned_set), key=str.lower)

def build_voucher_types_request(company=""):
    comp_var = f"<SVCOMPANYNAME>{escape_xml(company)}</SVCOMPANYNAME>" if company else "<SVCOMPANYNAME>$$CurrentCompany</SVCOMPANYNAME>"
    return f"""<ENVELOPE>
    <HEADER>
        <VERSION>1</VERSION>
        <TALLYREQUEST>Export</TALLYREQUEST>
        <TYPE>Collection</TYPE>
        <ID>AllVoucherTypes</ID>
    </HEADER>
    <BODY>
        <DESC>
            <STATICVARIABLES>
                {comp_var}
            </STATICVARIABLES>
            <TDL>
                <TDLMESSAGE>
                    <COLLECTION NAME="AllVoucherTypes" ISMODIFY="No">
                        <TYPE>VoucherType</TYPE>
                        <NATIVEMETHOD>Name</NATIVEMETHOD>
                        <NATIVEMETHOD>Parent</NATIVEMETHOD>
                    </COLLECTION>
                </TDLMESSAGE>
            </TDL>
        </DESC>
    </BODY>
</ENVELOPE>"""

def extract_voucher_types_from_xml(raw_xml):
    raw_names = re.findall(r'<VOUCHERTYPE\b[^>]*\bNAME="([^"]+)"', raw_xml, flags=re.IGNORECASE)
    if not raw_names:
        raw_names = re.findall(r"<NAME\b[^>]*>(.*?)</NAME>", raw_xml, flags=re.DOTALL | re.IGNORECASE)
    cleaned_set = set()
    for n in raw_names:
        clean_str = re.sub(r"<[^>]+>", "", n)
        clean_str = html.unescape(clean_str.strip())
        if clean_str and not clean_str.startswith("$$"):
            cleaned_set.add(clean_str)
    return sorted(list(cleaned_set), key=str.lower)

def build_daybook_request(company, from_date, to_date):
    return f"""<ENVELOPE>
  <HEADER>
    <VERSION>1</VERSION>
    <TALLYREQUEST>Export</TALLYREQUEST>
    <TYPE>Collection</TYPE>
    <ID>CustomVouchersFetch</ID>
  </HEADER>
  <BODY>
    <DESC>
      <STATICVARIABLES>
        <SVCURRENTCOMPANY>{escape_xml(company)}</SVCURRENTCOMPANY>
        <SVFROMDATE TYPE="Date">{from_date}</SVFROMDATE>
        <SVTODATE TYPE="Date">{to_date}</SVTODATE>
        <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
      </STATICVARIABLES>
      <TDL>
        <TDLMESSAGE>
          <COLLECTION NAME="CustomVouchersFetch" ISINITIALIZE="Yes">
            <TYPE>Voucher</TYPE>
            <FETCH>MASTERID, VOUCHERNUMBER, DATE, VOUCHERTYPENAME, NARRATION, PARTYLEDGERNAME, ALLLEDGERENTRIES.LEDGERNAME, ALLLEDGERENTRIES.AMOUNT, ALLLEDGERENTRIES.ISDEEMEDPOSITIVE, LEDGERENTRIES.LEDGERNAME, LEDGERENTRIES.AMOUNT, LEDGERENTRIES.ISDEEMEDPOSITIVE</FETCH>
          </COLLECTION>
        </TDLMESSAGE>
      </TDL>
    </DESC>
  </BODY>
</ENVELOPE>"""

def build_alter_request(company, vch_type, master_id, vch_number, date, narration, ledger_entries, party_ledger=""):
    entries_xml = ""
    for entry in ledger_entries:
        entries_xml += f"""
            <ALLLEDGERENTRIES.LIST>
                <LEDGERNAME>{escape_xml(entry['name'])}</LEDGERNAME>
                <ISDEEMEDPOSITIVE>{entry['deemed_positive']}</ISDEEMEDPOSITIVE>
                <AMOUNT>{entry['amount']}</AMOUNT>
            </ALLLEDGERENTRIES.LIST>"""
    party_line = f"<PARTYLEDGERNAME>{escape_xml(party_ledger)}</PARTYLEDGERNAME>" if party_ledger else ""
    return f"""<ENVELOPE>
  <HEADER>
    <TALLYREQUEST>Import Data</TALLYREQUEST>
  </HEADER>
  <BODY>
    <IMPORTDATA>
      <REQUESTDESC>
        <REPORTNAME>Vouchers</REPORTNAME>
        <STATICVARIABLES>
          <SVCURRENTCOMPANY>{escape_xml(company)}</SVCURRENTCOMPANY>
        </STATICVARIABLES>
      </REQUESTDESC>
      <REQUESTDATA>
        <TALLYMESSAGE xmlns:UDF="TallyUDF">
          <VOUCHER VCHTYPE="{escape_xml(vch_type)}" ACTION="Alter" ID="{master_id}">
            <VOUCHERNUMBER>{escape_xml(vch_number)}</VOUCHERNUMBER>
            <DATE>{date}</DATE>
            <VOUCHERTYPENAME>{escape_xml(vch_type)}</VOUCHERTYPENAME>
            {party_line}
            <NARRATION>{escape_xml(narration)}</NARRATION>
            {entries_xml}
          </VOUCHER>
        </TALLYMESSAGE>
      </REQUESTDATA>
    </IMPORTDATA>
  </BODY>
</ENVELOPE>"""

def build_create_voucher_request(company, vch_type, date, narration, ledger_entries, party_ledger=""):
    entries_xml = ""
    for entry in ledger_entries:
        entries_xml += f"""
            <ALLLEDGERENTRIES.LIST>
                <LEDGERNAME>{escape_xml(entry['name'])}</LEDGERNAME>
                <ISDEEMEDPOSITIVE>{entry['deemed_positive']}</ISDEEMEDPOSITIVE>
                <AMOUNT>{entry['amount']}</AMOUNT>
            </ALLLEDGERENTRIES.LIST>"""
    party_line = f"<PARTYLEDGERNAME>{escape_xml(party_ledger)}</PARTYLEDGERNAME>" if party_ledger else ""
    return f"""<ENVELOPE>
  <HEADER>
    <TALLYREQUEST>Import Data</TALLYREQUEST>
  </HEADER>
  <BODY>
    <IMPORTDATA>
      <REQUESTDESC>
        <REPORTNAME>Vouchers</REPORTNAME>
        <STATICVARIABLES>
          <SVCURRENTCOMPANY>{escape_xml(company)}</SVCURRENTCOMPANY>
        </STATICVARIABLES>
      </REQUESTDESC>
      <REQUESTDATA>
        <TALLYMESSAGE xmlns:UDF="TallyUDF">
          <VOUCHER VCHTYPE="{escape_xml(vch_type)}" ACTION="Create">
            <DATE>{date}</DATE>
            <VOUCHERTYPENAME>{escape_xml(vch_type)}</VOUCHERTYPENAME>
            {party_line}
            <NARRATION>{escape_xml(narration)}</NARRATION>
            {entries_xml}
          </VOUCHER>
        </TALLYMESSAGE>
      </REQUESTDATA>
    </IMPORTDATA>
  </BODY>
</ENVELOPE>"""

def extract_voucher_blocks(raw_xml):
    return re.findall(r"<VOUCHER\b[^>]*>.*?</VOUCHER>", raw_xml, flags=re.DOTALL)

def get_attr(block, attr):
    m = re.search(rf'{attr}="([^"]*)"', block)
    return m.group(1) if m else ""

def get_tag(block, tag):
    m = re.search(rf"<{tag}\b[^>]*>(.*?)</{tag}>", block, flags=re.DOTALL | re.IGNORECASE)
    if m:
        return html.unescape(m.group(1).strip())
    return ""

def get_ledger_names(block):
    return [
        html.unescape(name.strip())
        for name in re.findall(r"<LEDGERNAME\b[^>]*>(.*?)</LEDGERNAME>", block, flags=re.DOTALL | re.IGNORECASE)
    ]

def extract_ledger_entries(block):
    entries = []
    entry_blocks = re.findall(
        r"<(?:ALLLEDGERENTRIES|LEDGERENTRIES)\.LIST\b[^>]*>(.*?)</(?:ALLLEDGERENTRIES|LEDGERENTRIES)\.LIST>",
        block,
        flags=re.DOTALL | re.IGNORECASE,
    )
    for entry_block in entry_blocks:
        name = get_tag(entry_block, "LEDGERNAME")
        deemed = get_tag(entry_block, "ISDEEMEDPOSITIVE") or "No"
        raw_amt = get_tag(entry_block, "AMOUNT") or "0.00"
        amt_val = clean_float(raw_amt)
        if name:
            entries.append({"name": name, "deemed_positive": deemed, "amount": f"{amt_val:.2f}"})
    return entries

# ============================================================================
#  3. BACKGROUND THREAD WORKERS (Original + Create)
# ============================================================================
class TallyConnectWorker(QThread):
    finished = Signal(str, str, str, list, list)
    error = Signal(str)
    log_msg = Signal(str, str)
    def __init__(self, url):
        super().__init__()
        self.url = url

    def run(self):
        try:
            self.log_msg.emit(f"Connecting to Tally at {self.url}...", "INFO")
            xml_req = build_active_company_request()
            resp = requests.post(
                self.url,
                data=xml_req.encode("utf-8"),
                headers={"Content-Type": "text/xml"},
                timeout=8,
            )
            raw_xml = strip_invalid_xml_chars(resp.text)
            comp_name = get_tag(raw_xml, "NAME") or get_attr(raw_xml, "NAME")
            period_from = normalize_tally_date(get_tag(raw_xml, "PERIODFROM")) or "20250401"
            period_to = normalize_tally_date(get_tag(raw_xml, "PERIODTO")) or "20250430"
            self.log_msg.emit(f"Fetching ledger masters for '{comp_name or 'Current Company'}'...", "INFO")
            ledger_req = build_ledger_list_request(comp_name)
            ledger_resp = requests.post(
                self.url,
                data=ledger_req.encode("utf-8"),
                headers={"Content-Type": "text/xml"},
                timeout=10,
            )
            raw_ledger_xml = strip_invalid_xml_chars(ledger_resp.text)
            fetched_ledgers = extract_ledger_names_from_xml(raw_ledger_xml)
            self.log_msg.emit(f"Fetching dynamic voucher types from Tally XML Collection...", "INFO")
            vch_types_req = build_voucher_types_request(comp_name)
            vch_types_resp = requests.post(
                self.url,
                data=vch_types_req.encode("utf-8"),
                headers={"Content-Type": "text/xml"},
                timeout=10,
            )
            raw_vch_xml = strip_invalid_xml_chars(vch_types_resp.text)
            fetched_voucher_types = extract_voucher_types_from_xml(raw_vch_xml)
            self.finished.emit(comp_name, period_from, period_to, fetched_ledgers, fetched_voucher_types)
        except Exception as e:
            self.error.emit(str(e))

class TallyFetchWorker(QThread):
    progress = Signal(int, int, str)
    finished = Signal(list, float)
    error = Signal(str)
    log_msg = Signal(str, str)
    def __init__(self, url, company, from_date, to_date, ledger_filter=""):
        super().__init__()
        self.url = url
        self.company = company
        self.from_date = from_date
        self.to_date = to_date
        self.ledger_filter = ledger_filter.strip().lower()

    def run(self):
        start_time = time.perf_counter()
        date_chunks = generate_month_ranges(self.from_date, self.to_date)
        total_chunks = len(date_chunks)
        matched_blocks = []
        self.log_msg.emit(f"Starting chunked retrieval ({total_chunks} month slices) from {self.url}...", "INFO")
        for idx, (sub_from, sub_to) in enumerate(date_chunks, 1):
            self.progress.emit(idx, total_chunks, f"Fetching {sub_from} to {sub_to} ({idx}/{total_chunks})...")
            req_xml = build_daybook_request(self.company, sub_from, sub_to)
            try:
                resp = requests.post(
                    self.url,
                    data=req_xml.encode("utf-8"),
                    headers={"Content-Type": "text/xml"},
                    timeout=35,
                )
                raw_xml = strip_invalid_xml_chars(resp.text)
                blocks = extract_voucher_blocks(raw_xml)
                for b in blocks:
                    if not get_tag(b, "MASTERID").strip():
                        continue
                    if self.ledger_filter:
                        party = get_tag(b, "PARTYLEDGERNAME").lower()
                        ledgers = [l.lower() for l in get_ledger_names(b)]
                        if not (self.ledger_filter in party or any(self.ledger_filter in l for l in ledgers)):
                            continue
                    matched_blocks.append(b)
            except Exception as e:
                self.log_msg.emit(f"Chunk [{sub_from}-{sub_to}] failed: {e}", "WARN")
        elapsed = time.perf_counter() - start_time
        self.finished.emit(matched_blocks, elapsed)

class TallyPostWorker(QThread):
    progress = Signal(int, int, str)
    finished = Signal(int, int)
    log_msg = Signal(str, str)
    def __init__(self, url, company, tasks):
        super().__init__()
        self.url = url
        self.company = company
        self.tasks = tasks

    def run(self):
        total = len(self.tasks)
        success_count = 0
        for idx, item in enumerate(self.tasks, 1):
            master_id = item["master_id"]
            self.progress.emit(idx, total, f"Posting Master ID {master_id} ({idx}/{total})...")
            req_xml = build_alter_request(
                company=self.company,
                vch_type=item["vch_type"],
                master_id=master_id,
                vch_number=item["vch_no"],
                date=format_date_to_tally(item["date"]),
                narration=item["narration"],
                ledger_entries=item["ledger_entries"],
                party_ledger=item.get("party_ledger", ""),
            )
            try:
                resp = requests.post(
                    self.url,
                    data=req_xml.encode("utf-8"),
                    headers={"Content-Type": "text/xml"},
                    timeout=30,
                )
                if "<ALTERED>1</ALTERED>" in resp.text:
                    success_count += 1
                    self.log_msg.emit(f"Master ID {master_id} successfully updated (ALTERED=1).", "OK")
                else:
                    self.log_msg.emit(f"Master ID {master_id} response: {resp.text[:120]}", "WARN")
            except Exception as e:
                self.log_msg.emit(f"Master ID {master_id} failed: {e}", "ERR")
        self.finished.emit(success_count, total)

class TallyCreateWorker(QThread):
    progress = Signal(int, int, str)
    finished = Signal(int, int)
    log_msg = Signal(str, str)
    def __init__(self, url, company, tasks):
        super().__init__()
        self.url = url
        self.company = company
        self.tasks = tasks

    def run(self):
        total = len(self.tasks)
        success_count = 0
        for idx, item in enumerate(self.tasks, 1):
            self.progress.emit(idx, total, f"Creating voucher {idx}/{total}...")
            req_xml = build_create_voucher_request(
                company=self.company,
                vch_type=item["vch_type"],
                date=format_date_to_tally(item["date"]),
                narration=item["narration"],
                ledger_entries=item["ledger_entries"],
                party_ledger=item.get("party_ledger", ""),
            )
            try:
                resp = requests.post(
                    self.url,
                    data=req_xml.encode("utf-8"),
                    headers={"Content-Type": "text/xml"},
                    timeout=30,
                )
                if "<CREATED>1</CREATED>" in resp.text or "<ALTERED>1</ALTERED>" in resp.text:
                    success_count += 1
                    item["success"] = True
                    self.log_msg.emit(f"Voucher {idx} successfully created.", "OK")
                else:
                    self.log_msg.emit(f"Voucher {idx} response: {resp.text[:120]}", "WARN")
            except Exception as e:
                self.log_msg.emit(f"Voucher {idx} failed: {e}", "ERR")
        self.finished.emit(success_count, total)

# ============================================================================
#  4. EXCEL CUSTOM DIALOGS (Filter, Visibility, Mapping)
# ============================================================================
class CheckableListWidget(QListWidget):
    def mousePressEvent(self, event):
        pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
        item = self.itemAt(pos)
        if item and (item.flags() & Qt.ItemFlag.ItemIsUserCheckable):
            rect = self.visualItemRect(item)
            if pos.x() > rect.left() + 22:
                new_state = Qt.CheckState.Unchecked if item.checkState() == Qt.CheckState.Checked else Qt.CheckState.Checked
                item.setCheckState(new_state)
        super().mousePressEvent(event)

class ExcelColumnFilterDialog(QDialog):
    def __init__(self, parent, col_title, unique_values, current_selected=None, is_numeric=False):
        super().__init__(parent)
        self.setWindowTitle(f"Filter & Sort: {col_title}")
        self.setMinimumWidth(400 if is_numeric else 320)
        self.setMinimumHeight(420)
        self.unique_values = unique_values
        self.selected_values = set(current_selected if current_selected is not None else unique_values)
        self.sort_order = None
        self.is_numeric = is_numeric
        self.min_val = None
        self.max_val = None
        self.init_ui(col_title)

    def init_ui(self, col_title):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        title_lbl = QLabel(f"<b>Filter & Sort: {col_title}</b>")
        title_lbl.setStyleSheet("font-size: 13px; color: #0f3d33; font-family: 'Lexend', 'Segoe UI';")
        layout.addWidget(title_lbl)

        sort_layout = QHBoxLayout()
        btn_sort_asc = QPushButton("▲ Sort Smallest to Largest" if self.is_numeric else "▲ Sort A to Z")
        btn_sort_asc.setStyleSheet("font-family: 'Lexend'; font-size: 11px; padding: 4px;")
        btn_sort_asc.clicked.connect(lambda: self.set_sort_and_accept("ASC"))
        sort_layout.addWidget(btn_sort_asc)
        btn_sort_desc = QPushButton("▼ Sort Largest to Smallest" if self.is_numeric else "▼ Sort Z to A")
        btn_sort_desc.setStyleSheet("font-family: 'Lexend'; font-size: 11px; padding: 4px;")
        btn_sort_desc.clicked.connect(lambda: self.set_sort_and_accept("DESC"))
        sort_layout.addWidget(btn_sort_desc)
        layout.addLayout(sort_layout)

        if self.is_numeric:
            range_layout = QHBoxLayout()
            range_layout.addWidget(QLabel("Min:"))
            self.min_spin = QDoubleSpinBox()
            self.min_spin.setRange(-9999999, 9999999)
            self.min_spin.setValue(0)
            self.min_spin.setStyleSheet("font-family: 'Lexend'; font-size: 11px;")
            range_layout.addWidget(self.min_spin)
            range_layout.addWidget(QLabel("Max:"))
            self.max_spin = QDoubleSpinBox()
            self.max_spin.setRange(-9999999, 9999999)
            self.max_spin.setValue(9999999)
            self.max_spin.setStyleSheet("font-family: 'Lexend'; font-size: 11px;")
            range_layout.addWidget(self.max_spin)
            layout.addLayout(range_layout)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("🔍 Search values...")
        self.search_edit.setStyleSheet("border: 1px solid #c4bfae; padding: 4px 6px; font-family: 'Lexend'; font-size: 11px;")
        self.search_edit.textChanged.connect(self.filter_list_items)
        layout.addWidget(self.search_edit)

        self.chk_all = QCheckBox("(Select All)")
        self.chk_all.setChecked(len(self.selected_values) == len(self.unique_values))
        self.chk_all.setStyleSheet("font-family: 'Lexend'; font-weight: bold; color: #0f3d33;")
        self.chk_all.toggled.connect(self.toggle_all_items)
        layout.addWidget(self.chk_all)

        self.list_widget = CheckableListWidget()
        self.list_widget.setStyleSheet("""
            QListWidget {
                border: 1px solid #c4bfae;
                background-color: #ffffff;
                font-family: 'Lexend', 'Segoe UI';
                font-size: 11px;
            }
            QListWidget::item {
                padding: 3px 2px;
            }
        """)
        layout.addWidget(self.list_widget, 1)
        self.populate_list(self.unique_values)

        btn_layout = QHBoxLayout()
        btn_clear = QPushButton("Clear Filter")
        btn_clear.setStyleSheet("color: #b91c1c; font-family: 'Lexend'; font-size: 11px; padding: 4px 8px;")
        btn_clear.clicked.connect(self.clear_filter)
        btn_layout.addWidget(btn_clear)
        btn_layout.addStretch()
        btn_cancel = QPushButton("Cancel")
        btn_cancel.setStyleSheet("font-family: 'Lexend'; font-size: 11px; padding: 4px 10px;")
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(btn_cancel)
        btn_apply = QPushButton("Apply Filter")
        btn_apply.setStyleSheet("background-color: #107c41; color: #ffffff; font-weight: bold; font-family: 'Lexend'; font-size: 11px; padding: 4px 14px; border-radius: 3px;")
        btn_apply.clicked.connect(self.apply_filter)
        btn_layout.addWidget(btn_apply)
        layout.addLayout(btn_layout)

    def populate_list(self, values):
        self.list_widget.clear()
        for val in values:
            display_text = val if val.strip() else "<Blank>"
            item = QListWidgetItem(display_text, self.list_widget)
            item.setData(Qt.ItemDataRole.UserRole, val)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            check_state = Qt.CheckState.Checked if val in self.selected_values else Qt.CheckState.Unchecked
            item.setCheckState(check_state)

    def filter_list_items(self, text):
        query = text.strip().lower()
        matched = [v for v in self.unique_values if query in v.lower()]
        self.populate_list(matched)

    def toggle_all_items(self, checked):
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            item.setCheckState(Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked)

    def set_sort_and_accept(self, order):
        self.sort_order = order
        self.accept()

    def clear_filter(self):
        self.selected_values = None
        self.min_val = None
        self.max_val = None
        self.accept()

    def apply_filter(self):
        selected = set()
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                selected.add(item.data(Qt.ItemDataRole.UserRole))
        self.selected_values = selected
        if self.is_numeric:
            self.min_val = self.min_spin.value()
            self.max_val = self.max_spin.value()
        self.accept()

class ColumnVisibilityDialog(QDialog):
    def __init__(self, parent, column_headers, hidden_columns):
        super().__init__(parent)
        self.setWindowTitle("Hide & Unhide Columns")
        self.setMinimumWidth(340)
        self.setMinimumHeight(380)
        self.column_headers = column_headers
        self.hidden_columns = set(hidden_columns)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        lbl = QLabel("<b>Select columns to display in grid:</b>")
        lbl.setStyleSheet("font-family: 'Lexend', 'Segoe UI'; font-size: 12px; color: #0f3d33;")
        layout.addWidget(lbl)

        self.list_widget = CheckableListWidget()
        self.list_widget.setStyleSheet("border: 1px solid #c4bfae; font-family: 'Lexend', 'Segoe UI'; font-size: 11px;")
        layout.addWidget(self.list_widget, 1)

        for col_idx, (col_name, _) in enumerate(self.column_headers):
            item = QListWidgetItem(f"Column {col_idx}: {col_name}", self.list_widget)
            item.setData(Qt.ItemDataRole.UserRole, col_idx)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            is_checked = col_idx not in self.hidden_columns
            item.setCheckState(Qt.CheckState.Checked if is_checked else Qt.CheckState.Unchecked)

        btn_layout = QHBoxLayout()
        btn_show_all = QPushButton("Show All")
        btn_show_all.setStyleSheet("font-family: 'Lexend'; font-size: 11px; padding: 4px 8px;")
        btn_show_all.clicked.connect(self.show_all)
        btn_layout.addWidget(btn_show_all)
        btn_layout.addStretch()
        btn_done = QPushButton("Done")
        btn_done.setStyleSheet("background-color: #0f3d33; color: #ffffff; font-weight: bold; font-family: 'Lexend'; font-size: 11px; padding: 4px 14px; border-radius: 3px;")
        btn_done.clicked.connect(self.save_and_close)
        btn_layout.addWidget(btn_done)
        layout.addLayout(btn_layout)

    def show_all(self):
        for i in range(self.list_widget.count()):
            self.list_widget.item(i).setCheckState(Qt.CheckState.Checked)

    def save_and_close(self):
        new_hidden = set()
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            col_idx = item.data(Qt.ItemDataRole.UserRole)
            if item.checkState() != Qt.CheckState.Checked:
                new_hidden.add(col_idx)
        self.hidden_columns = new_hidden
        self.accept()

class ColumnMappingDialog(QDialog):
    def __init__(self, parent, headers, sample_data):
        super().__init__(parent)
        self.setWindowTitle("Map Columns to Fields")
        self.setMinimumWidth(500)
        self.headers = headers
        self.sample_data = sample_data
        self.mapping = {}
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)

        lbl = QLabel("Please select which column corresponds to each field:")
        lbl.setStyleSheet("font-family: 'Lexend'; font-size: 12px;")
        layout.addWidget(lbl)

        form = QFormLayout()
        form.setSpacing(6)

        self.combos = {}
        fields = ["Date", "Description", "Debit", "Credit", "Balance"]
        for field in fields:
            combo = QComboBox()
            combo.addItem("-- None --")
            for h in self.headers:
                combo.addItem(h)
            combo.setStyleSheet("font-family: 'Lexend';")
            self.combos[field] = combo
            for h in self.headers:
                hl = h.lower()
                if (field.lower() == "date" and ("date" in hl or "transaction date" in hl)) or \
                   (field.lower() == "description" and any(k in hl for k in ("description", "particulars", "narration"))) or \
                   (field.lower() == "debit" and any(k in hl for k in ("debit", "dr"))) or \
                   (field.lower() == "credit" and any(k in hl for k in ("credit", "cr"))) or \
                   (field.lower() == "balance" and "balance" in hl):
                    combo.setCurrentText(h)
                    break
            form.addRow(f"{field}:", combo)

        layout.addLayout(form)

        group = QGroupBox("Sample Data (first 3 rows)")
        group.setStyleSheet("font-family: 'Lexend';")
        sample_layout = QVBoxLayout()
        sample_text = QTextEdit()
        sample_text.setReadOnly(True)
        sample_text.setFixedHeight(120)
        sample_text.setStyleSheet("font-family: monospace; font-size: 10px;")
        lines = []
        for i, row in enumerate(self.sample_data[:3]):
            lines.append(f"Row {i+1}: " + " | ".join(str(cell) for cell in row))
        sample_text.setText("\n".join(lines))
        sample_layout.addWidget(sample_text)
        group.setLayout(sample_layout)
        layout.addWidget(group)

        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def get_mapping(self):
        mapping = {}
        for field, combo in self.combos.items():
            val = combo.currentText()
            if val != "-- None --":
                mapping[field] = val
        return mapping

# ============================================================================
#  5. EXCEL CUSTOM SPREADSHEET TABLE & DELEGATE – with dynamic columns support
# ============================================================================
class SearchableComboBox(QComboBox):
    def showPopup(self):
        if self.completer():
            self.completer().setCompletionPrefix(self.lineEdit().text())
            self.completer().complete()
        else:
            super().showPopup()

class ExcelTableDelegate(QStyledItemDelegate):
    def __init__(self, parent=None, ledgers=None, voucher_types=None):
        super().__init__(parent)
        self.ledgers = ledgers or []
        self.voucher_types = voucher_types or []

    def set_ledgers(self, ledgers):
        self.ledgers = ledgers

    def set_voucher_types(self, voucher_types):
        self.voucher_types = voucher_types

    def _find_parent_with_method(self, method_name):
        widget = self.parent()
        while widget:
            if hasattr(widget, method_name):
                return widget
            widget = widget.parent()
        return None

    def paint(self, painter, option, index):
        super().paint(painter, option, index)
        col = index.column()
        row = index.row()
        parent_widget = self._find_parent_with_method('get_ledger_columns')
        if parent_widget and hasattr(parent_widget, 'get_ledger_columns'):
            ledger_cols = parent_widget.get_ledger_columns()
            if row > 0 and col in ledger_cols:
                painter.save()
                painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
                arrow_w = 8
                arrow_h = 5
                x = option.rect.right() - 14
                y = option.rect.top() + (option.rect.height() - arrow_h) // 2
                painter.setPen(Qt.PenStyle.NoPen)
                painter.setBrush(QBrush(QColor("#6b8079")))
                points = [
                    QPoint(x, y),
                    QPoint(x + arrow_w, y),
                    QPoint(x + arrow_w // 2, y + arrow_h)
                ]
                painter.drawPolygon(QPolygon(points))
                painter.restore()
                return
        if row > 0 and col in (3, 8, 10, 12, 14):
            painter.save()
            painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
            arrow_w = 8
            arrow_h = 5
            x = option.rect.right() - 14
            y = option.rect.top() + (option.rect.height() - arrow_h) // 2
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QBrush(QColor("#6b8079")))
            points = [
                QPoint(x, y),
                QPoint(x + arrow_w, y),
                QPoint(x + arrow_w // 2, y + arrow_h)
            ]
            painter.drawPolygon(QPolygon(points))
            painter.restore()

    def createEditor(self, parent, option, index):
        row = index.row()
        col = index.column()
        if row == 0:
            return None
        if col == 2:  # Date
            date_edit = QDateEdit(parent)
            date_edit.setCalendarPopup(True)
            date_edit.setDisplayFormat("dd-MM-yyyy")
            date_edit.setStyleSheet("""
                QDateEdit {
                    border: 2px solid #107c41;
                    background-color: #ffffff;
                    color: #10221c;
                    font-family: 'Lexend', 'Segoe UI', sans-serif;
                    font-size: 11px;
                    padding: 1px 3px;
                }
                QDateEdit QCalendarWidget {
                    background-color: white;
                    color: black;
                }
            """)
            cal = date_edit.calendarWidget()
            if cal:
                cal.setStyleSheet("""
                    QCalendarWidget {
                        background-color: white;
                        color: black;
                    }
                    QCalendarWidget QAbstractItemView {
                        background-color: white;
                        color: black;
                    }
                    QCalendarWidget QAbstractItemView:selected {
                        background-color: #107c41;
                        color: white;
                    }
                    QCalendarWidget QToolButton {
                        color: black;
                        background-color: white;
                    }
                    QCalendarWidget QMenu {
                        color: black;
                        background-color: white;
                    }
                    QCalendarWidget QSpinBox {
                        color: black;
                        background-color: white;
                    }
                """)
            return date_edit
        if col == 3:  # VType
            combo = SearchableComboBox(parent)
            combo.setEditable(True)
            combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
            combo.setStyleSheet("""
                QComboBox {
                    border: 2px solid #107c41;
                    background-color: #ffffff;
                    color: #10221c;
                    font-family: 'Lexend', 'Segoe UI', sans-serif;
                    font-size: 11px;
                    padding: 1px 3px;
                }
                QComboBox::drop-down {
                    subcontrol-origin: padding;
                    subcontrol-position: top right;
                    width: 18px;
                    border-left: 1px solid #c4bfae;
                    background-color: #f3f1ea;
                }
                QComboBox::down-arrow {
                    image: none;
                    width: 0;
                    height: 0;
                    border-left: 4px solid transparent;
                    border-right: 4px solid transparent;
                    border-top: 5px solid #0f3d33;
                }
                QComboBox QAbstractItemView {
                    border: 1px solid #107c41;
                    background-color: #ffffff;
                    color: #10221c;
                    selection-background-color: #107c41;
                    selection-color: #ffffff;
                    font-family: 'Lexend', 'Segoe UI', sans-serif;
                    font-size: 11px;
                }
            """)
            combo.addItem("")
            for vt in self.voucher_types:
                combo.addItem(vt)
            completer = QCompleter(self.voucher_types, combo)
            completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
            completer.setFilterMode(Qt.MatchFlag.MatchContains)
            combo.setCompleter(completer)
            QTimer.singleShot(0, combo.showPopup)
            return combo
        parent_widget = self._find_parent_with_method('get_ledger_columns')
        if parent_widget and hasattr(parent_widget, 'get_ledger_columns'):
            ledger_cols = parent_widget.get_ledger_columns()
            amount_cols = parent_widget.get_amount_columns() if hasattr(parent_widget, 'get_amount_columns') else []
            if col in ledger_cols:
                combo = SearchableComboBox(parent)
                combo.setEditable(True)
                combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
                combo.setStyleSheet("""
                    QComboBox {
                        border: 2px solid #107c41;
                        background-color: #ffffff;
                        color: #10221c;
                        font-family: 'Lexend', 'Segoe UI', sans-serif;
                        font-size: 11px;
                        padding: 1px 3px;
                    }
                    QComboBox::drop-down {
                        subcontrol-origin: padding;
                        subcontrol-position: top right;
                        width: 18px;
                        border-left: 1px solid #c4bfae;
                        background-color: #f3f1ea;
                    }
                    QComboBox::down-arrow {
                        image: none;
                        width: 0;
                        height: 0;
                        border-left: 4px solid transparent;
                        border-right: 4px solid transparent;
                        border-top: 5px solid #0f3d33;
                    }
                    QComboBox QAbstractItemView {
                        border: 1px solid #107c41;
                        background-color: #ffffff;
                        color: #10221c;
                        selection-background-color: #107c41;
                        selection-color: #ffffff;
                        font-family: 'Lexend', 'Segoe UI', sans-serif;
                        font-size: 11px;
                    }
                """)
                combo.addItem("")
                for led in self.ledgers:
                    combo.addItem(led)
                completer = QCompleter(self.ledgers, combo)
                completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                completer.setFilterMode(Qt.MatchFlag.MatchContains)
                combo.setCompleter(completer)
                QTimer.singleShot(0, combo.showPopup)
                return combo
            if col in amount_cols:
                line_edit = QLineEdit(parent)
                line_edit.setStyleSheet("""
                    QLineEdit {
                        border: 2px solid #107c41;
                        background-color: #ffffff;
                        color: #10221c;
                        font-family: 'Lexend', 'Segoe UI', sans-serif;
                        font-size: 11px;
                        padding: 1px 4px;
                    }
                """)
                return line_edit
        if col == 7:  # Narration
            line_edit = QLineEdit(parent)
            line_edit.setStyleSheet("""
                QLineEdit {
                    border: 2px solid #107c41;
                    background-color: #ffffff;
                    color: #10221c;
                    font-family: 'Lexend', 'Segoe UI', sans-serif;
                    font-size: 11px;
                    padding: 1px 4px;
                }
            """)
            return line_edit
        return None

    def setEditorData(self, editor, index):
        val = index.model().data(index, Qt.ItemDataRole.EditRole)
        if isinstance(editor, QDateEdit):
            if val is not None:
                d = QDate.fromString(str(val), "dd-MM-yyyy")
                if not d.isValid():
                    d = QDate.fromString(str(val), "yyyyMMdd")
                if d.isValid():
                    editor.setDate(d)
        elif isinstance(editor, QComboBox):
            if val is not None:
                editor.setCurrentText(str(val))
                if editor.lineEdit():
                    editor.lineEdit().selectAll()
        elif isinstance(editor, QLineEdit):
            if val is not None:
                editor.setText(str(val))
                editor.deselect()
                editor.setCursorPosition(len(editor.text()))

    def setModelData(self, editor, model, index):
        if isinstance(editor, QDateEdit):
            model.setData(index, editor.date().toString("dd-MM-yyyy"), Qt.ItemDataRole.EditRole)
        elif isinstance(editor, QComboBox):
            txt = editor.currentText().strip()
            col = index.column()
            if col == 3:
                options = self.voucher_types
                if txt:
                    matched = next((opt for opt in options if opt.lower() == txt.lower()), None)
                    if matched:
                        model.setData(index, matched, Qt.ItemDataRole.EditRole)
                    else:
                        return
                else:
                    model.setData(index, "", Qt.ItemDataRole.EditRole)
            else:
                options = self.ledgers
                if txt:
                    matched = next((opt for opt in options if opt.lower() == txt.lower()), None)
                    if matched:
                        model.setData(index, matched, Qt.ItemDataRole.EditRole)
                    else:
                        return
                else:
                    model.setData(index, "", Qt.ItemDataRole.EditRole)
        elif isinstance(editor, QLineEdit):
            model.setData(index, editor.text().strip(), Qt.ItemDataRole.EditRole)

class ExcelGridWidget(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumSize(200, 200)
        self.setAlternatingRowColors(False)
        self.setShowGrid(True)
        self.setGridStyle(Qt.PenStyle.SolidLine)
        self.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.SelectedClicked
            | QAbstractItemView.EditTrigger.AnyKeyPressed
        )
        self.setStyleSheet("""
            QTableWidget {
                background-color: #ffffff;
                alternate-background-color: #faf9f4;
                gridline-color: #d8d2c1;
                font-family: 'Lexend', 'Segoe UI', 'Arial';
                font-size: 11px;
                color: #10221c;
                selection-background-color: #e4efe9;
                selection-color: #10221c;
                border: 1px solid #c4bfae;
            }
            QHeaderView::section {
                background-color: #f3f1ea;
                color: #0f3d33;
                font-family: 'Lexend', 'Segoe UI';
                font-weight: bold;
                font-size: 11px;
                border: 1px solid #d8d2c1;
                padding: 4px 6px;
            }
        """)
        self.auto_edit_columns = [2, 3, 8, 10, 12, 14]  # default, updated dynamically
        self._mouse_edit = False

    def set_auto_edit_columns(self, col_list):
        self.auto_edit_columns = col_list

    def get_current_editor(self):
        """Returns the current editor widget or None."""
        if hasattr(self, 'currentEditor'):
            return self.currentEditor()
        else:
            # Fallback: use focus widget if it's a child of the viewport
            fw = self.focusWidget()
            if fw and fw.parent() == self.viewport():
                return fw
            return None

    def _show_popup_if_combo(self):
        editor = self.get_current_editor()
        if isinstance(editor, QComboBox):
            editor.showPopup()

    def keyPressEvent(self, event):
        curr = self.currentIndex()
        if not curr.isValid():
            super().keyPressEvent(event)
            return

        row = curr.row()
        col = curr.column()
        key = event.key()

        # Alt+Down: open dropdown
        if event.modifiers() == Qt.KeyboardModifier.AltModifier and key == Qt.Key.Key_Down:
            if row > 0 and col in self.auto_edit_columns:
                self.edit(curr)
                QTimer.singleShot(10, self._show_popup_if_combo)
                event.accept()
                return

        # Handle Delete/Backspace key to clear cell content
        if key in (Qt.Key_Delete, Qt.Key_Backspace):
            if row > 0 and col in self.auto_edit_columns:
                model = self.model()
                if model:
                    index = model.index(row, col)
                    model.setData(index, "", Qt.ItemDataRole.EditRole)
                    event.accept()
                    return

        # If a regular (non-navigation) key is pressed, and we are not in edit mode,
        # start editing and pass the key to the editor
        if key not in (Qt.Key.Key_Up, Qt.Key.Key_Down, Qt.Key.Key_Left, Qt.Key.Key_Right,
                       Qt.Key.Key_Return, Qt.Key.Key_Enter, Qt.Key.Key_Escape,
                       Qt.Key.Key_F2, Qt.Key.Key_Tab) and event.text():
            if row > 0 and col in self.auto_edit_columns:
                self.edit(curr)
                editor = self.get_current_editor()
                if editor:
                    if isinstance(editor, QComboBox):
                        editor.setEditText(event.text())
                        editor.showPopup()
                    else:
                        editor.keyPressEvent(event)
                event.accept()
                return

        # If in editing state, let the editor handle the key (unless it's a navigation key we want to handle)
        if self.state() == QTableWidget.State.EditingState:
            if key in (Qt.Key.Key_Up, Qt.Key.Key_Down, Qt.Key.Key_Left, Qt.Key.Key_Right):
                editor = self.get_current_editor()
                if editor:
                    self.commitData(editor)
                    self.closeEditor(editor, QAbstractItemDelegate.EndEditHint.NoHint)
                    self.setFocus()
                    curr = self.currentIndex()
                    if not curr.isValid():
                        super().keyPressEvent(event)
                        return
                    row = curr.row()
                    col = curr.column()
                    if key == Qt.Key.Key_Up:
                        new_row = row - 1 if row > 1 else self.rowCount() - 1
                        self.setCurrentCell(new_row, col)
                        self.setFocus()
                        event.accept()
                        return
                    elif key == Qt.Key.Key_Down:
                        new_row = row + 1 if row < self.rowCount() - 1 else 1
                        self.setCurrentCell(new_row, col)
                        self.setFocus()
                        event.accept()
                        return
                    elif key == Qt.Key.Key_Left:
                        new_col = col - 1 if col > 0 else self.columnCount() - 1
                        self.setCurrentCell(row, new_col)
                        self.setFocus()
                        event.accept()
                        return
                    elif key == Qt.Key.Key_Right:
                        new_col = col + 1 if col < self.columnCount() - 1 else 0
                        self.setCurrentCell(row, new_col)
                        self.setFocus()
                        event.accept()
                        return
            super().keyPressEvent(event)
            return

        cell_widget = self.cellWidget(row, col)
        if isinstance(cell_widget, QComboBox) and cell_widget.view().isVisible():
            super().keyPressEvent(event)
            return

        if key == Qt.Key.Key_F2:
            if row > 0:
                self.edit(curr)
                event.accept()
                return

        if key in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            if self.state() == QTableWidget.State.EditingState:
                editor = self.get_current_editor()
                if editor:
                    self.commitData(editor)
                    self.closeEditor(editor, QAbstractItemDelegate.EndEditHint.NoHint)
            if event.modifiers() == Qt.KeyboardModifier.ShiftModifier:
                if row > 1:
                    self.setCurrentCell(row - 1, col)
            else:
                if row < self.rowCount() - 1:
                    self.setCurrentCell(row + 1, col)
            self.setFocus()
            event.accept()
            return

        if row == 0:
            super().keyPressEvent(event)
            return

        if key == Qt.Key.Key_Up:
            new_row = row - 1 if row > 1 else self.rowCount() - 1
            self.setCurrentCell(new_row, col)
            self.setFocus()
            event.accept()
            return
        if key == Qt.Key.Key_Down:
            new_row = row + 1 if row < self.rowCount() - 1 else 1
            self.setCurrentCell(new_row, col)
            self.setFocus()
            event.accept()
            return
        if key == Qt.Key.Key_Left:
            new_col = col - 1 if col > 0 else self.columnCount() - 1
            self.setCurrentCell(row, new_col)
            self.setFocus()
            event.accept()
            return
        if key == Qt.Key.Key_Right:
            new_col = col + 1 if col < self.columnCount() - 1 else 0
            self.setCurrentCell(row, new_col)
            self.setFocus()
            event.accept()
            return

        super().keyPressEvent(event)

    def mousePressEvent(self, event):
        pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
        index = self.indexAt(pos)
        if index.isValid() and index.row() > 0 and index.column() in self.auto_edit_columns:
            self._mouse_edit = True
            super().mousePressEvent(event)
            self.edit(index)
            self._mouse_edit = False
            return
        super().mousePressEvent(event)

    def paintEvent(self, event):
        super().paintEvent(event)
        curr = self.currentIndex()
        if curr.isValid() and curr.row() > 0:
            painter = QPainter(self.viewport())
            painter.setRenderHint(QPainter.RenderHint.Antialiasing, False)
            rect = self.visualRect(curr)
            pen = QPen(QColor("#107c41"), 2)
            painter.setPen(pen)
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawRect(rect.adjusted(1, 1, -1, -1))
            square_size = 5
            sq_rect = QRect(
                rect.right() - square_size + 1,
                rect.bottom() - square_size + 1,
                square_size,
                square_size
            )
            painter.fillRect(sq_rect, QColor("#107c41"))
            painter.end()

# ============================================================================
#  6. VOUCHER EDITOR WIDGET (TAB 1) – DYNAMIC LEDGER COLUMNS
# ============================================================================
class VoucherEditorWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_app = parent
        self.ledgers = []
        self.voucher_types = []
        self.vouchers_data = []
        self.column_filters = {}
        self.hidden_columns = set()
        self.posted_rows = set()
        self.debit_count = 0
        self.credit_count = 0
        self.fixed_column_count = 8
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        top_bar = QFrame()
        top_bar.setStyleSheet("""
            QFrame {
                background-color: #ffffff;
                border: 1px solid #d8d2c1;
                border-radius: 4px;
                padding: 4px 8px;
                font-family: 'Lexend', 'Segoe UI';
            }
        """)
        top_layout = QHBoxLayout(top_bar)
        top_layout.setContentsMargins(4, 4, 4, 4)
        top_layout.setSpacing(8)

        logo_lbl = QLabel("TV")
        logo_lbl.setStyleSheet("""
            background-color: #0f3d33;
            color: #ffffff;
            font-family: 'Lexend';
            font-weight: bold;
            font-size: 11px;
            padding: 3px 6px;
            border-radius: 3px;
        """)
        top_layout.addWidget(logo_lbl)

        lbl_url = QLabel("URL:")
        lbl_url.setStyleSheet("font-family: 'Lexend';")
        top_layout.addWidget(lbl_url)
        self.url_edit = QLineEdit(DEFAULT_TALLY_URL)
        self.url_edit.setFixedWidth(160)
        self.url_edit.setStyleSheet("border: 1px solid #c4bfae; padding: 3px 5px; font-family: 'Lexend'; font-size: 11px;")
        top_layout.addWidget(self.url_edit)

        self.btn_connect = QPushButton("🔌 Connect")
        self.btn_connect.setStyleSheet("""
            QPushButton {
                background-color: #0f3d33; color: #ffffff; font-family: 'Lexend'; font-weight: bold; font-size: 11px; padding: 4px 12px; border-radius: 3px; border: none;
            }
            QPushButton:hover { background-color: #165648; }
        """)
        self.btn_connect.clicked.connect(self.on_connect_clicked)
        top_layout.addWidget(self.btn_connect)
        
        self.btn_refresh = QPushButton("↻ Refresh")
        self.btn_refresh.setStyleSheet("""
            QPushButton {
                background-color: #1a5c82; color: #ffffff; font-family: 'Lexend'; font-weight: bold; font-size: 11px; padding: 4px 12px; border-radius: 3px; border: none;
            }
            QPushButton:hover { background-color: #12405c; }
        """)
        self.btn_refresh.clicked.connect(self.on_connect_clicked)
        top_layout.addWidget(self.btn_refresh)

        sep1 = QFrame()
        sep1.setFrameShape(QFrame.Shape.VLine)
        sep1.setStyleSheet("color: #d8d2c1;")
        top_layout.addWidget(sep1)

        lbl_comp = QLabel("Company:")
        lbl_comp.setStyleSheet("font-family: 'Lexend';")
        top_layout.addWidget(lbl_comp)
        self.comp_edit = QLineEdit()
        self.comp_edit.setReadOnly(True)
        self.comp_edit.setPlaceholderText("Auto-fetched on Connect")
        self.comp_edit.setFixedWidth(180)
        self.comp_edit.setStyleSheet("border: 1px solid #d8d2c1; background-color: #f0ede1; color: #10221c; padding: 3px 5px; font-family: 'Lexend'; font-size: 11px;")
        top_layout.addWidget(self.comp_edit)

        lbl_p = QLabel("Period:")
        lbl_p.setStyleSheet("font-family: 'Lexend';")
        top_layout.addWidget(lbl_p)
        self.from_date_edit = QDateEdit()
        self.from_date_edit.setCalendarPopup(True)
        self.from_date_edit.setDisplayFormat("dd-MM-yyyy")
        self.from_date_edit.setDate(QDate(2025, 4, 1))
        self.from_date_edit.setFixedWidth(95)
        self.from_date_edit.setStyleSheet("border: 1px solid #c4bfae; background-color: #ffffff; padding: 2px 4px; font-family: 'Lexend'; font-size: 11px;")
        top_layout.addWidget(self.from_date_edit)
        lbl_to = QLabel("to")
        lbl_to.setStyleSheet("font-family: 'Lexend';")
        top_layout.addWidget(lbl_to)
        self.to_date_edit = QDateEdit()
        self.to_date_edit.setCalendarPopup(True)
        self.to_date_edit.setDisplayFormat("dd-MM-yyyy")
        self.to_date_edit.setDate(QDate(2025, 4, 30))
        self.to_date_edit.setFixedWidth(95)
        self.to_date_edit.setStyleSheet("border: 1px solid #c4bfae; background-color: #ffffff; padding: 2px 4px; font-family: 'Lexend'; font-size: 11px;")
        top_layout.addWidget(self.to_date_edit)

        lbl_led = QLabel("Ledger:")
        lbl_led.setStyleSheet("font-family: 'Lexend';")
        top_layout.addWidget(lbl_led)
        self.ledger_filter_combo = QComboBox()
        self.ledger_filter_combo.setEditable(True)
        self.ledger_filter_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.ledger_filter_combo.setFixedWidth(180)
        self.ledger_filter_combo.addItem("")
        for led in self.ledgers:
            self.ledger_filter_combo.addItem(led)
        self.ledger_filter_combo.setCurrentText("")
        if self.ledger_filter_combo.lineEdit():
            self.ledger_filter_combo.lineEdit().setPlaceholderText("Search ledger (Blank = All)...")
        comp = QCompleter(self.ledgers, self.ledger_filter_combo)
        comp.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        comp.setFilterMode(Qt.MatchFlag.MatchContains)
        self.ledger_filter_combo.setCompleter(comp)
        self.ledger_filter_combo.setStyleSheet("border: 1px solid #c4bfae; padding: 2px 4px; font-family: 'Lexend'; font-size: 11px;")
        top_layout.addWidget(self.ledger_filter_combo)

        self.btn_fetch = QPushButton("🔄 Fetch")
        self.btn_fetch.setStyleSheet("""
            QPushButton {
                background-color: #107c41;
                color: #ffffff;
                font-family: 'Lexend';
                font-weight: bold;
                font-size: 11px;
                padding: 4px 12px;
                border-radius: 3px;
            }
            QPushButton:hover { background-color: #166534; }
        """)
        self.btn_fetch.clicked.connect(self.on_fetch_clicked)
        top_layout.addWidget(self.btn_fetch)

        btn_cols = QPushButton("👁 Columns...")
        btn_cols.setStyleSheet("""
            QPushButton {
                background-color: #f3f1ea;
                color: #0f3d33;
                border: 1px solid #c4bfae;
                font-family: 'Lexend';
                font-size: 11px;
                padding: 4px 8px;
                border-radius: 3px;
            }
            QPushButton:hover { background-color: #e5e0d3; }
        """)
        btn_cols.clicked.connect(self.open_column_manager)
        top_layout.addWidget(btn_cols)

        top_layout.addStretch()

        self.status_badge = QLabel("● Not Connected")
        self.status_badge.setStyleSheet("color: #b5820a; font-family: 'Lexend'; font-weight: bold; font-size: 11px;")
        top_layout.addWidget(self.status_badge)

        layout.addWidget(top_bar)

        

        legend_bar = QHBoxLayout()
        legend_bar.setContentsMargins(2, 0, 2, 0)
        self.summary_count_lbl = QLabel("0 of 0 vouchers displayed")
        self.summary_count_lbl.setStyleSheet("color: #3a5049; font-family: 'Lexend'; font-weight: bold; font-size: 11px;")
        legend_bar.addWidget(self.summary_count_lbl)
        legend_bar.addStretch()

        dr_dot = QLabel("■ Debit (Dr)")
        dr_dot.setStyleSheet("color: #1d4ed8; font-family: 'Lexend'; font-weight: bold; font-size: 11px;")
        legend_bar.addWidget(dr_dot)

        cr_dot = QLabel("■ Credit (Cr)")
        cr_dot.setStyleSheet("color: #c2540c; font-family: 'Lexend'; font-weight: bold; font-size: 11px; margin-left: 12px;")
        legend_bar.addWidget(cr_dot)

        self.chk_unbalanced_only = QCheckBox("☐ Unbalanced only")
        self.chk_unbalanced_only.setStyleSheet("color: #b91c1c; font-family: 'Lexend'; font-weight: bold; font-size: 11px; margin-left: 12px;")
        self.chk_unbalanced_only.toggled.connect(self.on_filter_changed)
        legend_bar.addWidget(self.chk_unbalanced_only)

        layout.addLayout(legend_bar)

        self.table = ExcelGridWidget(self)
        self.table.setColumnCount(0)
        self.table.horizontalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.horizontalHeader().customContextMenuRequested.connect(self.on_header_context_menu)
        self.table.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        # Connect cellChanged to update balance on amount edit
        self.table.cellChanged.connect(self.on_cell_data_modified)
        layout.addWidget(self.table, 1)

        bottom_bar = QFrame()
        bottom_bar.setStyleSheet("""
            QFrame {
                background-color: #0f3d33;
                border-radius: 4px;
                padding: 4px 8px;
                font-family: 'Lexend', 'Segoe UI';
            }
        """)
        bot_layout = QHBoxLayout(bottom_bar)
        bot_layout.setContentsMargins(6, 4, 6, 4)

        self.match_lbl = QLabel("Vouchers Visible: 0")
        self.match_lbl.setStyleSheet("color: #4fd1a5; font-family: 'Lexend'; font-weight: bold; font-size: 11px;")
        bot_layout.addWidget(self.match_lbl)

        btn_select_all = QPushButton("Select All")
        btn_select_all.setStyleSheet("background-color: #145c4b; color: #ffffff; font-family: 'Lexend'; font-size: 10px; padding: 3px 8px; border-radius: 2px;")
        btn_select_all.clicked.connect(self.select_all_rows)
        bot_layout.addWidget(btn_select_all)

        btn_deselect = QPushButton("Deselect All")
        btn_deselect.setStyleSheet("background-color: #145c4b; color: #ffffff; font-family: 'Lexend'; font-size: 10px; padding: 3px 8px; border-radius: 2px;")
        btn_deselect.clicked.connect(self.deselect_all_rows)
        bot_layout.addWidget(btn_deselect)

        bot_layout.addStretch()

        self.btn_export = QPushButton("💾 Export XML")
        self.btn_export.setStyleSheet("""
            QPushButton {
                background-color: #1a5c82; color: #ffffff; font-family: 'Lexend'; font-weight: bold; font-size: 11px; padding: 6px 16px; border-radius: 4px; border: none;
            }
            QPushButton:hover { background-color: #12405c; }
        """)
        self.btn_export.clicked.connect(self.on_export_clicked)
        bot_layout.addWidget(self.btn_export)

        self.btn_post = QPushButton("🚀 Post Selected Alterations to Tally")
        self.btn_post.setStyleSheet("""
            QPushButton {
                background-color: #a8461f;
                color: #ffffff;
                font-family: 'Lexend';
                font-weight: bold;
                font-size: 11px;
                padding: 5px 16px;
                border-radius: 3px;
            }
            QPushButton:hover { background-color: #8f3a17; }
        """)
        self.btn_post.clicked.connect(self.on_post_clicked)
        bot_layout.addWidget(self.btn_post)

        layout.addWidget(bottom_bar)

        log_header_layout = QHBoxLayout()
        log_title = QLabel("System Log")
        log_title.setStyleSheet("font-family: 'Lexend'; font-weight: bold; font-size: 11px;")
        log_header_layout.addWidget(log_title)
        
        self.btn_toggle_log = QPushButton("📌")
        self.btn_toggle_log.setToolTip("Toggle Log Visibility")
        self.btn_toggle_log.setStyleSheet("font-family: 'Lexend'; font-size: 12px; padding: 2px 5px; border: none; background: transparent;")
        self.btn_toggle_log.setCheckable(True)
        self.btn_toggle_log.clicked.connect(lambda checked: self.log_console.setVisible(not checked))
        log_header_layout.addStretch()
        log_header_layout.addWidget(self.btn_toggle_log)
        layout.addLayout(log_header_layout)

        self.log_console = QTextEdit()
        self.log_console.setReadOnly(True)
        self.log_console.setFixedHeight(75)
        self.log_console.setStyleSheet("""
            QTextEdit {
                background-color: #0d1a16;
                color: #c9e4d8;
                font-family: 'Lexend', monospace;
                font-size: 10px;
                border: 1px solid #145c4b;
                border-radius: 3px;
                padding: 4px;
            }
        """)
        layout.addWidget(self.log_console)

        self.log("Tally Voucher Editor initialized.", "INFO")
        self.setup_empty_table()

    def setup_empty_table(self):
        self.table.blockSignals(True)
        self.table.setColumnCount(self.fixed_column_count)
        headers = ["☑", "Sr.", "Date", "Vch Type", "Vch No", "Master ID", "Balance?", "Narration"]
        self.table.setHorizontalHeaderLabels(headers)
        widths = [36, 48, 85, 105, 80, 85, 95, 250]
        for i, w in enumerate(widths):
            self.table.setColumnWidth(i, w)
        self.table.setRowCount(1)
        self.setup_sub_header_row()
        self.table.blockSignals(False)

    def setup_dynamic_headers(self):
        total_cols = self.fixed_column_count + 2 * self.debit_count + 2 * self.credit_count
        self.table.setColumnCount(total_cols)
        headers = ["☑", "Sr.", "Date", "Vch Type", "Vch No", "Master ID", "Balance?", "Narration"]
        for i in range(1, self.debit_count + 1):
            headers.append(f"Debit Ledger {i}")
            headers.append(f"Debit Amt {i}")
        for i in range(1, self.credit_count + 1):
            headers.append(f"Credit Ledger {i}")
            headers.append(f"Credit Amt {i}")
        self.table.setHorizontalHeaderLabels(headers)
        for col_idx in range(self.fixed_column_count, total_cols):
            if (col_idx - self.fixed_column_count) % 2 == 0:
                self.table.setColumnWidth(col_idx, 180)
            else:
                self.table.setColumnWidth(col_idx, 105)
        ledger_cols = self.get_ledger_columns()
        amount_cols = self.get_amount_columns()
        # Make Vch No (col 4) editable, add to auto_edit_columns
        auto_cols = [2, 3, 4, 7] + ledger_cols + amount_cols
        self.table.set_auto_edit_columns(auto_cols)
        self.setup_sub_header_row()

    def get_ledger_columns(self):
        total_cols = self.table.columnCount()
        ledger_cols = []
        for i in range(self.fixed_column_count, total_cols):
            if (i - self.fixed_column_count) % 2 == 0:
                ledger_cols.append(i)
        return ledger_cols

    def get_amount_columns(self):
        total_cols = self.table.columnCount()
        amount_cols = []
        for i in range(self.fixed_column_count, total_cols):
            if (i - self.fixed_column_count) % 2 == 1:
                amount_cols.append(i)
        return amount_cols

    def setup_sub_header_row(self):
        if self.table.rowCount() == 0:
            self.table.setRowCount(1)
        for c in range(self.table.columnCount()):
            self.table.removeCellWidget(0, c)

        chk_all = QCheckBox()
        chk_all.setChecked(True)
        chk_all.setStyleSheet("margin-left: 10px;")
        chk_all.toggled.connect(lambda state: self.toggle_all_checkboxes(state))
        self.table.setCellWidget(0, 0, chk_all)

        for c in range(1, 3):
            lbl = QLabel("")
            lbl.setStyleSheet("background-color: #faf9f4;")
            self.table.setCellWidget(0, c, lbl)

        w_vtype = QWidget()
        w_vtype_layout = QHBoxLayout(w_vtype)
        w_vtype_layout.setContentsMargins(1, 1, 1, 1)
        w_vtype_layout.setSpacing(2)
        vtype_combo = SearchableComboBox()
        vtype_combo.setEditable(True)
        for vt in self.voucher_types:
            vtype_combo.addItem(vt)
        vtype_completer = QCompleter(self.voucher_types, vtype_combo)
        vtype_completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        vtype_completer.setFilterMode(Qt.MatchFlag.MatchContains)
        vtype_combo.setCompleter(vtype_completer)
        vtype_combo.setCurrentIndex(-1)
        vtype_combo.lineEdit().setPlaceholderText("-- Apply VType --")
        vtype_combo.setStyleSheet("font-family: 'Lexend'; font-size: 10px; border: 1px solid #c4bfae;")
        w_vtype_layout.addWidget(vtype_combo, 1)
        btn_vtype = QPushButton("⚡")
        btn_vtype.setFixedWidth(22)
        btn_vtype.setToolTip("Apply this VType to all checked rows")
        btn_vtype.setStyleSheet("background-color: #1d4ed8; color: #ffffff; font-family: 'Lexend'; font-weight: bold; font-size: 10px; padding: 1px; border: none; border-radius: 2px;")
        btn_vtype.clicked.connect(lambda: self.apply_bulk_vtype(vtype_combo.currentText()))
        w_vtype_layout.addWidget(btn_vtype)
        self.table.setCellWidget(0, 3, w_vtype)

        for c in range(4, 7):
            lbl = QLabel("")
            lbl.setStyleSheet("background-color: #faf9f4;")
            self.table.setCellWidget(0, c, lbl)

        self.narration_filter_edit = QLineEdit()
        self.narration_filter_edit.setPlaceholderText("🔍 Filter text...")
        self.narration_filter_edit.setStyleSheet("border: 1px solid #c4bfae; font-family: 'Lexend'; font-size: 10px; padding: 1px 3px; background-color: #ffffff;")
        self.narration_filter_edit.textChanged.connect(self.on_filter_changed)
        self.table.setCellWidget(0, 7, self.narration_filter_edit)

        ledger_cols = self.get_ledger_columns()
        for idx, col in enumerate(ledger_cols):
            if col < self.fixed_column_count + 2 * self.debit_count:
                prefix = "Dr"
                num = (col - self.fixed_column_count) // 2 + 1
            else:
                prefix = "Cr"
                num = (col - self.fixed_column_count - 2 * self.debit_count) // 2 + 1
            w = QWidget()
            w_layout = QHBoxLayout(w)
            w_layout.setContentsMargins(1, 1, 1, 1)
            w_layout.setSpacing(2)
            combo = SearchableComboBox()
            combo.setEditable(True)
            for led in self.ledgers:
                combo.addItem(led)
            ledger_completer = QCompleter(self.ledgers, combo)
            ledger_completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
            ledger_completer.setFilterMode(Qt.MatchFlag.MatchContains)
            combo.setCompleter(ledger_completer)
            combo.setCurrentIndex(-1)
            combo.lineEdit().setPlaceholderText(f"-- Apply {prefix}{num} --")
            combo.setStyleSheet("font-family: 'Lexend'; font-size: 10px; border: 1px solid #c4bfae;")
            w_layout.addWidget(combo, 1)
            btn = QPushButton("⚡")
            btn.setFixedWidth(22)
            btn.setToolTip(f"Apply this ledger to all checked rows for {prefix}{num}")
            btn.setStyleSheet("background-color: #1d4ed8; color: #ffffff; font-family: 'Lexend'; font-weight: bold; font-size: 10px; padding: 1px; border: none; border-radius: 2px;")
            btn.clicked.connect(lambda _, c=col, combo=combo: self.apply_bulk_ledger(c, combo.currentText()))
            w_layout.addWidget(btn)
            self.table.setCellWidget(0, col, w)

        for col in self.get_amount_columns():
            lbl = QLabel("")
            lbl.setStyleSheet("background-color: #faf9f4;")
            self.table.setCellWidget(0, col, lbl)

        for c in range(self.table.columnCount()):
            it = self.table.item(0, c)
            if it:
                it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)

    def apply_bulk_vtype(self, vtype_name):
        if not vtype_name or vtype_name.startswith("--"):
            return
        applied = 0
        for r in range(1, self.table.rowCount()):
            chk = self.table.cellWidget(r, 0)
            if isinstance(chk, QCheckBox) and chk.isChecked() and not self.table.isRowHidden(r):
                item = self.table.item(r, 3)
                if item:
                    item.setText(vtype_name)
                    applied += 1
        self.log(f"Applied VType '{vtype_name}' to {applied} rows.", "OK")

    def apply_bulk_ledger(self, col_idx, ledger_name):
        if not ledger_name or ledger_name.startswith("--"):
            return
        applied = 0
        for r in range(1, self.table.rowCount()):
            chk = self.table.cellWidget(r, 0)
            if isinstance(chk, QCheckBox) and chk.isChecked() and not self.table.isRowHidden(r):
                item = self.table.item(r, col_idx)
                if item:
                    item.setText(ledger_name)
                    applied += 1
        self.log(f"Applied '{ledger_name}' to {applied} selected rows.", "OK")

    def log(self, text, level="INFO"):
        ts = datetime.now().strftime("%H:%M:%S")
        color = "#c9e4d8"
        if level == "OK":
            color = "#4fd1a5"
        elif level == "WARN":
            color = "#e8b84b"
        elif level == "ERR":
            color = "#f08a7a"
        self.log_console.append(f"<span style='color:#5d7a70;'>[{ts}]</span> <span style='color:{color}; font-family: Lexend;'>{html.escape(text)}</span>")

    def on_connect_clicked(self):
        url = self.url_edit.text().strip()
        if not url:
            QMessageBox.warning(self, "Missing URL", "Please enter a valid Tally Prime URL.")
            return

        self.btn_connect.setEnabled(False)
        self.status_badge.setText("● Connecting...")
        self.status_badge.setStyleSheet("color: #b5820a; font-family: 'Lexend'; font-weight: bold; font-size: 11px;")

        self.conn_worker = TallyConnectWorker(url)
        self.conn_worker.finished.connect(self.on_connect_success)
        self.conn_worker.error.connect(self.on_connect_error)
        self.conn_worker.log_msg.connect(self.log)
        self.conn_worker.start()

    def on_connect_success(self, comp_name, from_date, to_date, ledgers, voucher_types):
        self.btn_connect.setEnabled(True)
        self.parent_app.active_company = comp_name
        self.parent_app.ledgers = ledgers
        self.parent_app.voucher_types = voucher_types

        self.comp_edit.setText(comp_name)

        d_from = QDate.fromString(from_date, "yyyyMMdd")
        if d_from.isValid():
            self.from_date_edit.setDate(d_from)
        d_to = QDate.fromString(to_date, "yyyyMMdd")
        if d_to.isValid():
            self.to_date_edit.setDate(d_to)

        self.ledgers = ledgers
        self.voucher_types = voucher_types
        self.delegate = ExcelTableDelegate(self.table, self.ledgers, self.voucher_types)
        self.table.setItemDelegate(self.delegate)

        self.ledger_filter_combo.clear()
        self.ledger_filter_combo.addItem("")
        for l in ledgers:
            self.ledger_filter_combo.addItem(l)
        self.ledger_filter_combo.setCurrentText("")
        if self.ledger_filter_combo.lineEdit():
            self.ledger_filter_combo.lineEdit().setPlaceholderText("Search ledger (Blank = All)...")
        comp = QCompleter(ledgers, self.ledger_filter_combo)
        comp.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        comp.setFilterMode(Qt.MatchFlag.MatchContains)
        self.ledger_filter_combo.setCompleter(comp)

        self.status_badge.setText(f"● Connected ({comp_name or 'Tally Prime'})")
        self.status_badge.setStyleSheet("color: #4fd1a5; font-family: 'Lexend'; font-weight: bold; font-size: 11px;")

        self.parent_app.enable_bank_import(True)

        self.log(f"Connected to Tally Prime. Loaded {len(ledgers)} ledger masters and {len(voucher_types)} dynamic voucher types.", "OK")

    def on_connect_error(self, err):
        self.btn_connect.setEnabled(True)
        self.status_badge.setText("● Disconnected")
        self.status_badge.setStyleSheet("color: #f08a7a; font-family: 'Lexend'; font-weight: bold; font-size: 11px;")
        self.parent_app.enable_bank_import(False)
        self.log(f"Connection failed: {err}", "ERR")
        QMessageBox.warning(self, "Connection Error", f"Could not reach Tally on {self.url_edit.text()}\n\nDetails: {err}")

    def on_fetch_clicked(self):
        url = self.url_edit.text().strip()
        comp = self.comp_edit.text().strip()
        f_date = self.from_date_edit.date().toString("yyyyMMdd")
        t_date = self.to_date_edit.date().toString("yyyyMMdd")
        raw_led_filter = self.ledger_filter_combo.currentText().strip()
        matched_led = next((l for l in self.ledgers if l.lower() == raw_led_filter.lower()), "")
        led_filter = matched_led if raw_led_filter else ""

        if not comp:
            QMessageBox.warning(self, "Missing Company", "Please enter or connect to a company first.")
            return

        self.btn_fetch.setEnabled(False)
        self.fetch_worker = TallyFetchWorker(url, comp, f_date, t_date, led_filter)
        self.fetch_worker.progress.connect(lambda cur, tot, msg: self.log(msg, "INFO"))
        self.fetch_worker.finished.connect(self.on_fetch_success)
        self.fetch_worker.log_msg.connect(self.log)
        self.fetch_worker.start()

    def on_fetch_success(self, raw_blocks, elapsed):
        self.btn_fetch.setEnabled(True)
        self.vouchers_data = raw_blocks
        self.log(f"Fetched {len(raw_blocks)} voucher(s) in {elapsed:.2f} seconds.", "OK")
        max_debits = 0
        max_credits = 0
        for block in raw_blocks:
            entries = extract_ledger_entries(block)
            debits = [e for e in entries if e["deemed_positive"].upper() in ("YES", "TRUE", "1")]
            credits = [e for e in entries if e["deemed_positive"].upper() in ("NO", "FALSE", "0")]
            max_debits = max(max_debits, len(debits))
            max_credits = max(max_credits, len(credits))
        self.debit_count = max_debits
        self.credit_count = max_credits
        self.setup_dynamic_headers()
        self.populate_table(raw_blocks)

    def populate_table(self, blocks):
        self.table.blockSignals(True)
        total_rows = len(blocks) + 1
        self.table.setRowCount(total_rows)
        self.setup_sub_header_row()
        self.posted_rows.clear()

        for idx, block in enumerate(blocks, 1):
            entries = extract_ledger_entries(block)
            debits = [e for e in entries if e["deemed_positive"].upper() in ("YES", "TRUE", "1")]
            credits = [e for e in entries if e["deemed_positive"].upper() in ("NO", "FALSE", "0")]
            tot_d = sum(clean_float(e["amount"]) for e in debits)
            tot_c = sum(clean_float(e["amount"]) for e in credits)
            diff = round(abs(tot_d - tot_c), 2)
            is_balanced = (diff == 0.0)

            chk = QCheckBox()
            chk.setChecked(True)
            chk.setStyleSheet("margin-left: 10px;")
            self.table.setCellWidget(idx, 0, chk)

            sr_item = QTableWidgetItem(str(idx))
            sr_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            sr_item.setForeground(QBrush(QColor("#a8461f")))
            sr_item.setFont(QFont("Lexend", 9, QFont.Weight.Bold))
            sr_item.setFlags(sr_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(idx, 1, sr_item)

            dt_item = QTableWidgetItem(format_date_to_ui(get_tag(block, "DATE")))
            dt_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(idx, 2, dt_item)

            vtype = get_attr(block, "VCHTYPE") or "Payment"
            v_item = QTableWidgetItem(vtype)
            v_item.setForeground(QBrush(QColor("#145c4b")))
            v_item.setFont(QFont("Lexend", 9, QFont.Weight.Bold))
            self.table.setItem(idx, 3, v_item)

            # Vch No - make editable (no extra flags)
            vno_item = QTableWidgetItem(get_tag(block, "VOUCHERNUMBER"))
            vno_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(idx, 4, vno_item)

            mid_item = QTableWidgetItem(get_tag(block, "MASTERID"))
            mid_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            mid_item.setForeground(QBrush(QColor("#6b8079")))
            mid_item.setFlags(mid_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(idx, 5, mid_item)

            bal_item = QTableWidgetItem()
            bal_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            bal_item.setFont(QFont("Lexend", 9, QFont.Weight.Bold))
            bal_item.setFlags(bal_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            if is_balanced:
                bal_item.setText("Balanced")
                bal_item.setForeground(QBrush(QColor("#107c41")))
            else:
                bal_item.setText(f"Diff: {diff:.2f}")
                bal_item.setForeground(QBrush(QColor("#b91c1c")))
            self.table.setItem(idx, 6, bal_item)

            narr_item = QTableWidgetItem(get_tag(block, "NARRATION"))
            self.table.setItem(idx, 7, narr_item)

            for i in range(self.debit_count):
                ledger_col = self.fixed_column_count + 2 * i
                amt_col = ledger_col + 1
                if i < len(debits):
                    debit = debits[i]
                    led_item = QTableWidgetItem(debit["name"])
                    led_item.setForeground(QBrush(QColor("#1d4ed8")))
                    led_item.setFont(QFont("Lexend", 9, QFont.Weight.Bold))
                    self.table.setItem(idx, ledger_col, led_item)
                    amt_item = QTableWidgetItem(debit["amount"])
                    amt_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    amt_item.setForeground(QBrush(QColor("#1d4ed8")))
                    amt_item.setFont(QFont("Lexend", 9, QFont.Weight.Bold))
                    self.table.setItem(idx, amt_col, amt_item)
                else:
                    led_item = QTableWidgetItem("")
                    led_item.setForeground(QBrush(QColor("#1d4ed8")))
                    self.table.setItem(idx, ledger_col, led_item)
                    amt_item = QTableWidgetItem("")
                    amt_item.setForeground(QBrush(QColor("#1d4ed8")))
                    self.table.setItem(idx, amt_col, amt_item)

            credit_start = self.fixed_column_count + 2 * self.debit_count
            for i in range(self.credit_count):
                ledger_col = credit_start + 2 * i
                amt_col = ledger_col + 1
                if i < len(credits):
                    credit = credits[i]
                    led_item = QTableWidgetItem(credit["name"])
                    led_item.setForeground(QBrush(QColor("#c2540c")))
                    led_item.setFont(QFont("Lexend", 9, QFont.Weight.Bold))
                    self.table.setItem(idx, ledger_col, led_item)
                    amt_item = QTableWidgetItem(credit["amount"])
                    amt_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    amt_item.setForeground(QBrush(QColor("#c2540c")))
                    amt_item.setFont(QFont("Lexend", 9, QFont.Weight.Bold))
                    self.table.setItem(idx, amt_col, amt_item)
                else:
                    led_item = QTableWidgetItem("")
                    led_item.setForeground(QBrush(QColor("#c2540c")))
                    self.table.setItem(idx, ledger_col, led_item)
                    amt_item = QTableWidgetItem("")
                    amt_item.setForeground(QBrush(QColor("#c2540c")))
                    self.table.setItem(idx, amt_col, amt_item)

            bg_color = QColor("#ffffff" if idx % 2 == 0 else "#faf9f4")
            if not is_balanced:
                bg_color = QColor("#ffe6e6")
            for c in range(self.table.columnCount()):
                it = self.table.item(idx, c)
                if it:
                    it.setBackground(QBrush(bg_color))

        self.table.blockSignals(False)
        self.apply_all_row_filters()

    def update_row_balance(self, row):
        total_debit = 0.0
        total_credit = 0.0
        amount_cols = self.get_amount_columns()
        for amt_col in amount_cols:
            amt_str = self.get_cell_text(row, amt_col)
            amt = clean_float(amt_str)
            if amt > 0:
                if amt_col < self.fixed_column_count + 2 * self.debit_count:
                    total_debit += amt
                else:
                    total_credit += amt
        diff = round(abs(total_debit - total_credit), 2)
        is_bal = (diff == 0.0)
        bal_item = self.table.item(row, 6)
        if bal_item:
            if is_bal:
                bal_item.setText("Balanced")
                bal_item.setForeground(QBrush(QColor("#107c41")))
            else:
                bal_item.setText(f"Diff: {diff:.2f}")
                bal_item.setForeground(QBrush(QColor("#b91c1c")))
        bg_color = QColor("#ffffff" if row % 2 == 0 else "#faf9f4")
        if not is_bal:
            bg_color = QColor("#ffe6e6")
        for c in range(self.table.columnCount()):
            it = self.table.item(row, c)
            if not it:
                it = QTableWidgetItem("")
                self.table.setItem(row, c, it)
            it.setBackground(QBrush(bg_color))
        if row in self.posted_rows:
            self.posted_rows.discard(row)

    def on_cell_data_modified(self, row, col):
        if row <= 0:
            return
        amount_cols = self.get_amount_columns()
        if col in amount_cols:
            self.update_row_balance(row)

    def on_header_clicked(self, col_idx):
        if col_idx > 0:
            self.open_excel_column_filter(col_idx)

    def on_header_context_menu(self, pos):
        col_idx = self.table.horizontalHeader().logicalIndexAt(pos)
        if col_idx < 0:
            return
        col_name = self.table.horizontalHeaderItem(col_idx).text()
        menu = QMenu(self)
        menu.setStyleSheet("font-family: 'Lexend', 'Segoe UI'; font-size: 11px;")
        act_filter = menu.addAction(f"🔍 Filter & Sort '{col_name}'...")
        act_filter.triggered.connect(lambda: self.open_excel_column_filter(col_idx))
        menu.addSeparator()
        act_hide = menu.addAction(f"👁 Hide Column '{col_name}'")
        act_hide.triggered.connect(lambda: self.hide_column(col_idx))
        act_unhide_all = menu.addAction("👁 Unhide All Columns")
        act_unhide_all.triggered.connect(self.unhide_all_columns)
        menu.addSeparator()
        act_manage = menu.addAction("⚙ Column Visibility Manager...")
        act_manage.triggered.connect(self.open_column_manager)
        if col_idx in self.column_filters:
            act_clear_filter = menu.addAction(f"Clear Filter on '{col_name}'")
            act_clear_filter.triggered.connect(lambda: self.clear_column_filter(col_idx))
        menu.exec(self.table.horizontalHeader().mapToGlobal(pos))

    def hide_column(self, col_idx):
        self.hidden_columns.add(col_idx)
        self.table.setColumnHidden(col_idx, True)
        self.log(f"Hidden column {self.table.horizontalHeaderItem(col_idx).text()}.", "INFO")

    def unhide_all_columns(self):
        self.hidden_columns.clear()
        for c in range(self.table.columnCount()):
            self.table.setColumnHidden(c, False)
        self.log("All columns unhidden.", "OK")

    def open_column_manager(self):
        column_headers = [(self.table.horizontalHeaderItem(i).text(), self.table.columnWidth(i)) for i in range(self.table.columnCount())]
        dlg = ColumnVisibilityDialog(self, column_headers, self.hidden_columns)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.hidden_columns = dlg.hidden_columns
            for c in range(self.table.columnCount()):
                self.table.setColumnHidden(c, c in self.hidden_columns)

    def open_excel_column_filter(self, col_idx):
        col_title = self.table.horizontalHeaderItem(col_idx).text()
        unique_vals = set()
        is_numeric = col_idx in self.get_amount_columns()
        for r in range(1, self.table.rowCount()):
            it = self.table.item(r, col_idx)
            val = it.text() if it else ""
            unique_vals.add(val)
        sorted_vals = sorted(list(unique_vals), key=str.lower)
        curr_selected = self.column_filters.get(col_idx, None)
        dlg = ExcelColumnFilterDialog(self, col_title, sorted_vals, curr_selected, is_numeric)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            if dlg.sort_order:
                self.sort_data_rows(col_idx, dlg.sort_order)
            if dlg.selected_values is None:
                self.clear_column_filter(col_idx)
            else:
                if is_numeric and dlg.min_val is not None and dlg.max_val is not None:
                    self.column_filters[col_idx] = {"type": "range", "min": dlg.min_val, "max": dlg.max_val}
                else:
                    self.column_filters[col_idx] = {"type": "values", "values": dlg.selected_values}
                self.update_header_labels()
                self.apply_all_row_filters()

    def sort_data_rows(self, col_idx, order):
        if self.table.rowCount() <= 1:
            return
        rows_data = []
        for r in range(1, self.table.rowCount()):
            chk_widget = self.table.cellWidget(r, 0)
            is_checked = chk_widget.isChecked() if isinstance(chk_widget, QCheckBox) else True
            row_items = []
            for c in range(self.table.columnCount()):
                it = self.table.item(r, c)
                row_items.append(it.text() if it else "")
            v_block = self.vouchers_data[r - 1] if (r - 1 < len(self.vouchers_data)) else ""
            rows_data.append({
                "checked": is_checked,
                "items": row_items,
                "raw_block": v_block
            })
        if not rows_data:
            return

        def get_sort_key(row_dict):
            val = row_dict["items"][col_idx] if col_idx < len(row_dict["items"]) else ""
            if col_idx == 2 and val:
                parts = val.split("-")
                if len(parts) == 3:
                    return (0, f"{parts[2]}{parts[1]}{parts[0]}")
            if col_idx in self.get_amount_columns():
                cleaned = re.sub(r"[^d.-]", "", val)
                if cleaned:
                    try:
                        return (0, float(cleaned))
                    except ValueError:
                        pass
            if col_idx in (4, 5):
                cleaned = re.sub(r"[^d.-]", "", val)
                if cleaned:
                    try:
                        return (0, float(cleaned))
                    except ValueError:
                        pass
            return (1, val.lower())

        reverse = (order == "DESC")
        rows_data.sort(key=get_sort_key, reverse=reverse)

        self.table.blockSignals(True)
        for new_idx, r_data in enumerate(rows_data, start=1):
            if new_idx - 1 < len(self.vouchers_data):
                self.vouchers_data[new_idx - 1] = r_data["raw_block"]
            chk = self.table.cellWidget(new_idx, 0)
            if isinstance(chk, QCheckBox):
                chk.blockSignals(True)
                chk.setChecked(r_data["checked"])
                chk.blockSignals(False)
            for c_idx, text_val in enumerate(r_data["items"]):
                it = self.table.item(new_idx, c_idx)
                if not it:
                    it = QTableWidgetItem(text_val)
                    self.table.setItem(new_idx, c_idx, it)
                else:
                    it.setText(text_val)
            self.update_row_balance(new_idx)
        self.table.blockSignals(False)
        self.apply_all_row_filters()
        self.log(f"Sorted by column {col_idx} ({order}).", "INFO")

    def clear_column_filter(self, col_idx):
        if col_idx in self.column_filters:
            del self.column_filters[col_idx]
            self.update_header_labels()
            self.apply_all_row_filters()

    def update_header_labels(self):
        for col_idx in range(self.table.columnCount()):
            item = self.table.horizontalHeaderItem(col_idx)
            if item:
                orig_text = item.text().replace(" [▼]", "")
                if col_idx in self.column_filters:
                    item.setText(f"{orig_text} [▼]")
                else:
                    item.setText(orig_text)

    def apply_all_row_filters(self):
        unbal_only = self.chk_unbalanced_only.isChecked()
        narr_filter = self.narration_filter_edit.text().strip().lower() if hasattr(self, 'narration_filter_edit') else ""
        visible_count = 0
        for r in range(1, self.table.rowCount()):
            row_visible = True
            if unbal_only:
                bal_item = self.table.item(r, 6)
                if bal_item and bal_item.text() == "Balanced":
                    row_visible = False
            if row_visible and narr_filter:
                narr_item = self.table.item(r, 7)
                narr_text = narr_item.text().lower() if narr_item else ""
                if narr_filter not in narr_text:
                    row_visible = False
            if row_visible:
                for col_idx, filter_info in self.column_filters.items():
                    it = self.table.item(r, col_idx)
                    val = it.text() if it else ""
                    if filter_info.get("type") == "range":
                        try:
                            num_val = float(val.replace(',', '')) if val else 0.0
                            if not (filter_info["min"] <= num_val <= filter_info["max"]):
                                row_visible = False
                                break
                        except:
                            row_visible = False
                            break
                    else:
                        if val not in filter_info.get("values", set()):
                            row_visible = False
                            break
            self.table.setRowHidden(r, not row_visible)
            if row_visible:
                visible_count += 1
        total = max(0, self.table.rowCount() - 1)
        self.summary_count_lbl.setText(f"{visible_count} of {total} vouchers displayed")
        self.match_lbl.setText(f"Vouchers Visible: {visible_count}")

    def on_cell_changed(self, row, col, prev_row, prev_col):
        pass

    def on_formula_text_edited(self, text):
        pass



    def on_filter_changed(self):
        self.apply_all_row_filters()

    def toggle_all_checkboxes(self, state):
        for r in range(1, self.table.rowCount()):
            if not self.table.isRowHidden(r):
                w = self.table.cellWidget(r, 0)
                if isinstance(w, QCheckBox):
                    w.setChecked(state)

    def select_all_rows(self):
        self.toggle_all_checkboxes(True)

    def deselect_all_rows(self):
        self.toggle_all_checkboxes(False)

    def get_cell_text(self, row, col):
        it = self.table.item(row, col)
        return it.text().strip() if it else ""

    def get_selected_tasks(self):
        tasks = []
        for r in range(1, self.table.rowCount()):
            chk = self.table.cellWidget(r, 0)
            if isinstance(chk, QCheckBox) and chk.isChecked() and not self.table.isRowHidden(r):
                master_id = self.get_cell_text(r, 5)
                vch_no = self.get_cell_text(r, 4)
                date_str = self.get_cell_text(r, 2)
                vch_type = self.get_cell_text(r, 3)
                narr = self.get_cell_text(r, 7)

                entries = []
                for i in range(self.debit_count):
                    ledger_col = self.fixed_column_count + 2 * i
                    amt_col = ledger_col + 1
                    led_name = self.get_cell_text(r, ledger_col)
                    amt = clean_float(self.get_cell_text(r, amt_col))
                    if led_name and amt > 0:
                        entries.append({"name": led_name, "deemed_positive": "Yes", "amount": f"-{amt:.2f}"})
                credit_start = self.fixed_column_count + 2 * self.debit_count
                for i in range(self.credit_count):
                    ledger_col = credit_start + 2 * i
                    amt_col = ledger_col + 1
                    led_name = self.get_cell_text(r, ledger_col)
                    amt = clean_float(self.get_cell_text(r, amt_col))
                    if led_name and amt > 0:
                        entries.append({"name": led_name, "deemed_positive": "No", "amount": f"{amt:.2f}"})

                if master_id and entries:
                    tasks.append({
                        "master_id": master_id,
                        "vch_no": vch_no,
                        "date": date_str,
                        "vch_type": vch_type,
                        "narration": narr,
                        "ledger_entries": entries,
                        "party_ledger": entries[0]["name"] if entries else ""
                    })
        return tasks

    def on_export_clicked(self):
        comp = self.comp_edit.text().strip()
        if not comp:
            QMessageBox.warning(self, "Missing Company", "Please connect to a Tally company first.")
            return

        tasks = self.get_selected_tasks()
        if not tasks:
            QMessageBox.information(self, "No Tasks", "No checked rows with valid entries found to export.")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Save XML", "", "XML Files (*.xml)")
        if not file_path:
            return

        xml_content = "<ENVELOPE>\n<BODY>\n<IMPORTDATA>\n<REQUESTDATA>\n<TALLYMESSAGE xmlns:UDF=\"TallyUDF\">\n"
        for t in tasks:
            req = build_alter_request(comp, t['vch_type'], t['master_id'], t['vch_no'], format_date_to_tally(t['date']), t['narration'], t['ledger_entries'], t.get('party_ledger', ''))
            m = re.search(r'(<VOUCHER .*?</VOUCHER>)', req, flags=re.DOTALL)
            if m:
                xml_content += m.group(1) + "\n"
        xml_content += "</TALLYMESSAGE>\n</REQUESTDATA>\n</IMPORTDATA>\n</BODY>\n</ENVELOPE>"
        
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(xml_content)
        self.log(f"Exported {len(tasks)} vouchers to {file_path}", "OK")
        QMessageBox.information(self, "Export Complete", f"Successfully exported {len(tasks)} vouchers to XML.")

    def on_post_clicked(self):
        comp = self.comp_edit.text().strip()
        if not comp:
            QMessageBox.warning(self, "Missing Company", "Please connect to a Tally company first.")
            return

        tasks = self.get_selected_tasks()
        if not tasks:
            QMessageBox.information(self, "No Tasks", "No checked rows with valid entries found to post.")
            return

        reply = QMessageBox.question(
            self,
            "Confirm Batch Post",
            f"Are you sure you want to post {len(tasks)} voucher alterations directly to '{comp}' in Tally?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        self.btn_post.setEnabled(False)
        self.post_worker = TallyPostWorker(self.url_edit.text().strip(), comp, tasks)
        self.post_worker.progress.connect(lambda cur, tot, msg: self.log(msg, "INFO"))
        self.post_worker.finished.connect(self.on_post_finished)
        self.post_worker.log_msg.connect(self.log)
        self.post_worker.start()

    def on_post_finished(self, success_count, total):
        self.btn_post.setEnabled(True)
        self.log(f"Batch posting complete: {success_count}/{total} vouchers updated successfully.", "OK")

# ============================================================================
#  7. BANK IMPORT WIDGET (TAB 2) – with Amount 1 as balancing figure
# ============================================================================
class BankImportWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_app = parent
        self.statement_data = []
        self.mapping = {}
        self.bank_ledger = None
        self.pair_count = 1
        self.ledger_names = []
        self.voucher_types = []
        self._loading = False
        self.column_filters = {}
        self.hidden_columns = set()
        self.posted_rows = set()
        self._balancing = False
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        top_frame = QFrame()
        top_frame.setStyleSheet("background-color: #faf9f4; border-radius: 3px; padding: 6px;")
        top_layout = QHBoxLayout(top_frame)
        top_layout.setContentsMargins(6, 6, 6, 6)

        lbl_bank = QLabel("Bank Ledger:")
        lbl_bank.setStyleSheet("font-family: 'Lexend'; font-weight: bold;")
        top_layout.addWidget(lbl_bank)

        self.bank_combo = SearchableComboBox()
        self.bank_combo.setEditable(True)
        self.bank_combo.setCurrentIndex(-1)
        self.bank_combo.lineEdit().setPlaceholderText("-- Select Bank Ledger --")
        self.bank_combo.setStyleSheet("""
            QComboBox {
                border: 2px solid #107c41;
                background-color: #ffffff;
                color: #10221c;
                font-family: 'Lexend', 'Segoe UI', sans-serif;
                font-size: 11px;
                padding: 1px 3px;
                min-width: 200px;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 18px;
                border-left: 1px solid #c4bfae;
                background-color: #f3f1ea;
            }
        """)
        self.bank_combo.setMaxVisibleItems(20)
        self.bank_combo.view().setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.bank_combo.currentTextChanged.connect(self.on_bank_ledger_changed)
        top_layout.addWidget(self.bank_combo)

        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText("Select Excel file...")
        self.file_path_edit.setReadOnly(True)
        self.file_path_edit.setStyleSheet("font-family: 'Lexend'; font-size: 11px;")
        top_layout.addWidget(self.file_path_edit, 1)

        btn_load = QPushButton("📂 Load Excel")
        btn_load.setStyleSheet("font-family: 'Lexend'; font-weight: bold; padding: 4px 10px;")
        btn_load.clicked.connect(self.load_excel)
        top_layout.addWidget(btn_load)

        btn_template = QPushButton("📥 Download Template")
        btn_template.setStyleSheet("font-family: 'Lexend'; font-weight: bold; padding: 4px 10px;")
        btn_template.clicked.connect(self.download_template)
        top_layout.addWidget(btn_template)

        btn_clear = QPushButton("Clear")
        btn_clear.setStyleSheet("font-family: 'Lexend'; padding: 4px 10px;")
        btn_clear.clicked.connect(self.clear_data)
        top_layout.addWidget(btn_clear)

        layout.addWidget(top_frame)

        self.table = ExcelGridWidget(self)
        self.table.setColumnCount(7 + 2 * self.pair_count)
        self.table.horizontalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.horizontalHeader().customContextMenuRequested.connect(self.on_header_context_menu)
        self.table.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        self.table.cellChanged.connect(self.on_cell_changed)
        layout.addWidget(self.table, 1)

        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        btn_add = QPushButton("➕ Add Ledger Column")
        btn_add.setStyleSheet("font-family: 'Lexend'; font-weight: bold; padding: 4px 10px;")
        btn_add.clicked.connect(self.add_ledger_column)
        btn_layout.addWidget(btn_add)

        btn_remove = QPushButton("➖ Remove Last Ledger")
        btn_remove.setStyleSheet("font-family: 'Lexend'; padding: 4px 10px;")
        btn_remove.clicked.connect(self.remove_ledger_column)
        btn_layout.addWidget(btn_remove)
        
        self.btn_add_row = QPushButton("➕ Add Empty Row")
        self.btn_add_row.setStyleSheet("font-family: 'Lexend'; font-weight: bold; padding: 4px 10px; color: #107c41;")
        self.btn_add_row.clicked.connect(self.add_empty_row)
        btn_layout.addWidget(self.btn_add_row)

        btn_layout.addStretch()

        self.btn_validate = QPushButton("✅ Validate")
        self.btn_validate.setStyleSheet("background-color: #f39c12; color: white; font-weight: bold; padding: 4px 12px; border-radius: 3px;")
        self.btn_validate.clicked.connect(self.validate_data)
        btn_layout.addWidget(self.btn_validate)

        layout.addLayout(btn_layout)

        bottom_frame = QFrame()
        bottom_frame.setStyleSheet("background-color: #faf9f4; border-radius: 3px; padding: 4px;")
        bottom_layout = QHBoxLayout(bottom_frame)
        bottom_layout.setContentsMargins(6, 4, 6, 4)

        self.count_lbl = QLabel("0 rows loaded")
        self.count_lbl.setStyleSheet("font-family: 'Lexend'; font-weight: bold;")
        bottom_layout.addWidget(self.count_lbl)

        bottom_layout.addStretch()

        btn_select_all = QPushButton("Select All")
        btn_select_all.setStyleSheet("font-family: 'Lexend'; padding: 2px 8px;")
        btn_select_all.clicked.connect(self.select_all)
        bottom_layout.addWidget(btn_select_all)

        btn_deselect_all = QPushButton("Deselect All")
        btn_deselect_all.setStyleSheet("font-family: 'Lexend'; padding: 2px 8px;")
        btn_deselect_all.clicked.connect(self.deselect_all)
        bottom_layout.addWidget(btn_deselect_all)

        bottom_layout.addStretch()

        self.btn_manual_save = QPushButton("💾 Manual Save")
        self.btn_manual_save.setStyleSheet("background-color: #0f3d33; color: white; font-weight: bold; padding: 4px 12px; border-radius: 3px;")
        self.btn_manual_save.clicked.connect(self.save_to_excel)
        bottom_layout.addWidget(self.btn_manual_save)

        self.btn_export = QPushButton("💾 Export XML")
        self.btn_export.setStyleSheet("""
            QPushButton {
                background-color: #1a5c82; color: #ffffff; font-family: 'Lexend'; font-weight: bold; font-size: 11px; padding: 6px 16px; border-radius: 4px; border: none;
            }
            QPushButton:hover { background-color: #12405c; }
        """)
        self.btn_export.clicked.connect(self.on_export_clicked)
        bottom_layout.addWidget(self.btn_export)

        self.btn_post = QPushButton("🚀 Post Selected to Tally")
        self.btn_post.setStyleSheet("""
            QPushButton {
                background-color: #a8461f;
                color: #ffffff;
                font-family: 'Lexend';
                font-weight: bold;
                font-size: 11px;
                padding: 5px 16px;
                border-radius: 3px;
            }
            QPushButton:hover { background-color: #8f3a17; }
        """)
        self.btn_post.clicked.connect(self.post_selected)
        bottom_layout.addWidget(self.btn_post)

        layout.addWidget(bottom_frame)

        # Log console (like Voucher Editor)
        log_header_layout_2 = QHBoxLayout()
        log_title_2 = QLabel("System Log")
        log_title_2.setStyleSheet("font-family: 'Lexend'; font-weight: bold; font-size: 11px;")
        log_header_layout_2.addWidget(log_title_2)
        
        self.btn_toggle_log_2 = QPushButton("📌")
        self.btn_toggle_log_2.setToolTip("Toggle Log Visibility")
        self.btn_toggle_log_2.setStyleSheet("font-family: 'Lexend'; font-size: 12px; padding: 2px 5px; border: none; background: transparent;")
        self.btn_toggle_log_2.setCheckable(True)
        self.btn_toggle_log_2.clicked.connect(lambda checked: self.log_console.setVisible(not checked))
        log_header_layout_2.addStretch()
        log_header_layout_2.addWidget(self.btn_toggle_log_2)
        layout.addLayout(log_header_layout_2)

        self.log_console = QTextEdit()
        self.log_console.setReadOnly(True)
        self.log_console.setFixedHeight(75)
        self.log_console.setStyleSheet("""
            QTextEdit {
                background-color: #0d1a16;
                color: #c9e4d8;
                font-family: 'Lexend', monospace;
                font-size: 10px;
                border: 1px solid #145c4b;
                border-radius: 3px;
                padding: 4px;
            }
        """)
        layout.addWidget(self.log_console)
        self.log("Bank Import tab initialized.", "INFO")

        self.setEnabled(False)
        self.setup_table()

    def log(self, text, level="INFO"):
        ts = datetime.now().strftime("%H:%M:%S")
        color = "#c9e4d8"
        if level == "OK":
            color = "#4fd1a5"
        elif level == "WARN":
            color = "#e8b84b"
        elif level == "ERR":
            color = "#f08a7a"
        self.log_console.append(f"<span style='color:#5d7a70;'>[{ts}]</span> <span style='color:{color}; font-family: Lexend;'>{html.escape(text)}</span>")

    def setup_table(self):
        self.table.blockSignals(True)
        base_cols = ["☑", "Sr.", "Date", "Vch No.", "Narration", "Debit", "Credit", "VType"]
        pair_cols = []
        for i in range(1, self.pair_count + 1):
            pair_cols.append(f"Ledger {i}")
            pair_cols.append(f"Amount {i}")
        all_cols = base_cols + pair_cols
        self.table.setColumnCount(len(all_cols))
        self.table.setHorizontalHeaderLabels(all_cols)

        self.table.setColumnWidth(0, 40)
        self.table.setColumnWidth(1, 45)
        self.table.setColumnWidth(2, 100)
        self.table.setColumnWidth(3, 80)
        self.table.setColumnWidth(4, 200)
        self.table.setColumnWidth(5, 100)
        self.table.setColumnWidth(6, 100)
        self.table.setColumnWidth(7, 80)
        for i in range(8, self.table.columnCount()):
            if (i - 8) % 2 == 0:
                self.table.setColumnWidth(i, 180)
            else:
                self.table.setColumnWidth(i, 100)

        self.delegate = BankImportDelegate(self.table, self.parent_app.ledgers, self.parent_app.voucher_types)
        self.table.setItemDelegate(self.delegate)

        self.ledger_col_indices = [i for i in range(8, self.table.columnCount(), 2)]
        self.amount_col_indices = [i for i in range(9, self.table.columnCount(), 2)]

        auto_cols = [2, 3, 4, 5, 6, 7] + self.ledger_col_indices + self.amount_col_indices
        self.table.set_auto_edit_columns(auto_cols)

        if self.table.rowCount() == 0:
            self.table.setRowCount(1)

        for c in range(self.table.columnCount()):
            self.table.removeCellWidget(0, c)

        chk_all = QCheckBox()
        chk_all.setChecked(True)
        chk_all.setStyleSheet("margin-left: 10px;")
        chk_all.toggled.connect(lambda state: self.toggle_all_checkboxes(state))
        self.table.setCellWidget(0, 0, chk_all)

        for c in range(1, 4):
            lbl = QLabel("")
            lbl.setStyleSheet("background-color: #faf9f4;")
            self.table.setCellWidget(0, c, lbl)

        self.narration_filter_edit = QLineEdit()
        self.narration_filter_edit.setPlaceholderText("🔍 Filter text...")
        self.narration_filter_edit.setStyleSheet("border: 1px solid #c4bfae; font-family: 'Lexend'; font-size: 10px; padding: 1px 3px; background-color: #ffffff;")
        self.narration_filter_edit.textChanged.connect(self.on_narration_filter_changed)
        self.table.setCellWidget(0, 4, self.narration_filter_edit)

        for c in range(5, 7):
            lbl = QLabel("")
            lbl.setStyleSheet("background-color: #faf9f4;")
            self.table.setCellWidget(0, c, lbl)

        w_vtype = QWidget()
        w_vtype_layout = QHBoxLayout(w_vtype)
        w_vtype_layout.setContentsMargins(1, 1, 1, 1)
        w_vtype_layout.setSpacing(2)
        vtype_combo = SearchableComboBox()
        vtype_combo.setEditable(True)
        for vt in self.voucher_types:
            vtype_combo.addItem(vt)
        vtype_completer = QCompleter(self.voucher_types, vtype_combo)
        vtype_completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        vtype_completer.setFilterMode(Qt.MatchFlag.MatchContains)
        vtype_combo.setCompleter(vtype_completer)
        vtype_combo.setCurrentIndex(-1)
        vtype_combo.lineEdit().setPlaceholderText("-- Apply VType --")
        vtype_combo.setStyleSheet("font-family: 'Lexend'; font-size: 10px; border: 1px solid #c4bfae;")
        w_vtype_layout.addWidget(vtype_combo, 1)
        btn_vtype = QPushButton("⚡")
        btn_vtype.setFixedWidth(22)
        btn_vtype.setToolTip("Apply this VType to all checked rows")
        btn_vtype.setStyleSheet("background-color: #1d4ed8; color: #ffffff; font-family: 'Lexend'; font-weight: bold; font-size: 10px; padding: 1px; border: none; border-radius: 2px;")
        btn_vtype.clicked.connect(lambda: self.apply_bulk_vtype(vtype_combo.currentText()))
        w_vtype_layout.addWidget(btn_vtype)
        self.table.setCellWidget(0, 7, w_vtype)

        for i, ledger_col in enumerate(self.ledger_col_indices, start=1):
            w = QWidget()
            w_layout = QHBoxLayout(w)
            w_layout.setContentsMargins(1, 1, 1, 1)
            w_layout.setSpacing(2)
            combo = SearchableComboBox()
            combo.setEditable(True)
            for led in self.parent_app.ledgers:
                combo.addItem(led)
            ledger_completer = QCompleter(self.parent_app.ledgers, combo)
            ledger_completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
            ledger_completer.setFilterMode(Qt.MatchFlag.MatchContains)
            combo.setCompleter(ledger_completer)
            combo.setCurrentIndex(-1)
            combo.lineEdit().setPlaceholderText(f"-- Apply Ledger {i} --")
            combo.setStyleSheet("font-family: 'Lexend'; font-size: 10px; border: 1px solid #c4bfae;")
            w_layout.addWidget(combo, 1)
            btn = QPushButton("⚡")
            btn.setFixedWidth(22)
            btn.setToolTip(f"Apply this ledger to all checked rows for Ledger {i}")
            btn.setStyleSheet("background-color: #1d4ed8; color: #ffffff; font-family: 'Lexend'; font-weight: bold; font-size: 10px; padding: 1px; border: none; border-radius: 2px;")
            btn.clicked.connect(lambda _, c=ledger_col, combo=combo: self.apply_bulk_ledger(c, combo.currentText()))
            w_layout.addWidget(btn)
            self.table.setCellWidget(0, ledger_col, w)

        for c in self.amount_col_indices:
            lbl = QLabel("")
            lbl.setStyleSheet("background-color: #faf9f4;")
            self.table.setCellWidget(0, c, lbl)

        for c in range(self.table.columnCount()):
            it = self.table.item(0, c)
            if it:
                it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)

        self.table.blockSignals(False)

    def apply_bulk_vtype(self, vtype_name):
        if not vtype_name or vtype_name.startswith("--"):
            return
        applied = 0
        for r in range(1, self.table.rowCount()):
            chk = self.table.cellWidget(r, 0)
            if isinstance(chk, QCheckBox) and chk.isChecked() and not self.table.isRowHidden(r):
                item = self.table.item(r, 6)
                if item:
                    item.setText(vtype_name)
                    applied += 1
        QMessageBox.information(self, "Bulk Apply", f"Applied VType '{vtype_name}' to {applied} rows.")

    def apply_bulk_ledger(self, col_idx, ledger_name):
        if not ledger_name or ledger_name.startswith("--"):
            return
        applied = 0
        for r in range(1, self.table.rowCount()):
            chk = self.table.cellWidget(r, 0)
            if isinstance(chk, QCheckBox) and chk.isChecked() and not self.table.isRowHidden(r):
                item = self.table.item(r, col_idx)
                if item:
                    item.setText(ledger_name)
                    applied += 1
        QMessageBox.information(self, "Bulk Apply", f"Applied '{ledger_name}' to {applied} checked rows.")

    def on_narration_filter_changed(self, text):
        search_text = text.strip().lower()
        for r in range(1, self.table.rowCount()):
            item = self.table.item(r, 3)
            if item and search_text in item.text().lower():
                self.table.setRowHidden(r, False)
            else:
                self.table.setRowHidden(r, True)

    def toggle_all_checkboxes(self, state):
        for r in range(1, self.table.rowCount()):
            if not self.table.isRowHidden(r):
                w = self.table.cellWidget(r, 0)
                if isinstance(w, QCheckBox):
                    w.setChecked(state)

    def download_template(self):
        if Workbook is None:
            QMessageBox.critical(self, "Missing Dependency", "openpyxl is not installed. Please install it:\npip install openpyxl")
            return
        try:
            from openpyxl.styles import PatternFill, Font
        except ImportError:
            PatternFill = None
            Font = None

        base_headers = ["Transaction Date*", "Vch No.", "Narration", "Debit*", "Credit*"]
        pair_headers = []
        for i in range(1, self.pair_count + 1):
            pair_headers.append(f"Counter Ledger {i}")
            pair_headers.append(f"Counter Ledger {i} Amt")
        all_headers = base_headers + pair_headers
        wb = Workbook()
        ws = wb.active
        ws.title = "Bank Statement Template"
        
        header_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") if PatternFill else None
        header_font = Font(bold=True) if Font else None

        for col_idx, header in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            if header_fill:
                cell.fill = header_fill
            if header_font:
                cell.font = header_font
        sample1 = ["01-04-2025", "VCH-01", "Sample UPI payment", 10000, "", "A/C Payee", 10000, "", ""]
        sample2 = ["02-04-2025", "VCH-02", "Sample NEFT credit", "", 25000, "Customer A", 25000, "", ""]
        for row_idx, sample in enumerate([sample1, sample2], 2):
            for col_idx, val in enumerate(sample, 1):
                if col_idx <= len(all_headers):
                    ws.cell(row=row_idx, column=col_idx, value=val)
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[col_letter].width = adjusted_width
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Template",
            "Bank_Statement_Template.xlsx",
            "Excel Files (*.xlsx)"
        )
        if file_path:
            try:
                wb.save(file_path)
                QMessageBox.information(self, "Success", f"Template saved to:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Could not save template:\n{str(e)}")

    def add_ledger_column(self):
        self._save_current_data()
        self.pair_count += 1
        self.setup_table()
        for row_data in self.statement_data:
            row_data["ledger_entries"].append({"ledger": "", "amount": 0.0})
        self.populate_table()
        self.narration_filter_edit.clear()
        self.on_narration_filter_changed("")

    def remove_ledger_column(self):
        if self.pair_count > 1:
            self._save_current_data()
            self.pair_count -= 1
            self.setup_table()
            for row_data in self.statement_data:
                if row_data["ledger_entries"]:
                    row_data["ledger_entries"].pop()
            self.populate_table()
            self.narration_filter_edit.clear()
            self.on_narration_filter_changed("")

    def add_empty_row(self):
        r = self.table.rowCount()
        self.table.insertRow(r)
        
        chk = QCheckBox()
        chk.setChecked(True)
        chk.setStyleSheet("margin-left: 10px;")
        self.table.setCellWidget(r, 0, chk)
        
        sr_item = QTableWidgetItem(str(r))
        sr_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        sr_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(r, 1, sr_item)
        
        for c in range(2, self.table.columnCount()):
            it = QTableWidgetItem("")
            self.table.setItem(r, c, it)
        
        self.table.scrollToBottom()
        self.log(f"Added empty row {r}", "INFO")

    def _save_current_data(self):
        if self.table.rowCount() <= 1:
            return
        if self.statement_data is None:
            self.statement_data = []
        for r in range(1, self.table.rowCount()):
            idx = r - 1
            if idx >= len(self.statement_data):
                self.statement_data.append({
                    "date": "", "narration": "", "withdrawal": 0.0, "deposit": 0.0,
                    "vtype": "Payment", "ledger_entries": [], "checked": True, "raw_block": {}
                })
            row_data = self.statement_data[idx]
            
            d_item = self.table.item(r, 2)
            row_data["date"] = d_item.text() if d_item else ""
            
            n_item = self.table.item(r, 3)
            row_data["narration"] = n_item.text() if n_item else ""
            
            w_item = self.table.item(r, 4)
            row_data["withdrawal"] = clean_float(w_item.text()) if w_item else 0.0
            
            dep_item = self.table.item(r, 5)
            row_data["deposit"] = clean_float(dep_item.text()) if dep_item else 0.0
            
            vtype_item = self.table.item(r, 6)
            row_data["vtype"] = vtype_item.text() if vtype_item else "Payment"
            
            entries = []
            for pair_idx in range(self.pair_count):
                ledger_col = 7 + pair_idx * 2
                amount_col = 8 + pair_idx * 2
                led_item = self.table.item(r, ledger_col)
                amt_item = self.table.item(r, amount_col)
                led_name = led_item.text() if led_item else ""
                amt_val = clean_float(amt_item.text()) if amt_item else 0.0
                entries.append({"ledger": led_name, "amount": amt_val})
            row_data["ledger_entries"] = entries
            
            chk = self.table.cellWidget(r, 0)
            row_data["checked"] = chk.isChecked() if isinstance(chk, QCheckBox) else True

    def set_ledgers(self, ledgers):
        self.ledger_names = ledgers
        self.bank_combo.clear()
        for led in ledgers:
            self.bank_combo.addItem(led)
        completer = QCompleter(ledgers, self.bank_combo)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        self.bank_combo.setCompleter(completer)
        self.bank_combo.setCurrentIndex(-1)
        self.bank_combo.lineEdit().setPlaceholderText("-- Select Bank Ledger --")
        if hasattr(self, 'delegate'):
            self.delegate.set_ledgers(ledgers)
        self.setup_table()
        if self.statement_data:
            self.populate_table()

    def set_voucher_types(self, voucher_types):
        self.voucher_types = voucher_types
        if hasattr(self, 'delegate'):
            self.delegate.set_voucher_types(voucher_types)
        self.setup_table()
        if self.statement_data:
            self.populate_table()

    def on_bank_ledger_changed(self, text):
        self.bank_ledger = text if text != "-- Select Bank Ledger --" else None

    def load_excel(self):
        if openpyxl is None:
            QMessageBox.critical(self, "Missing Dependency", "openpyxl is not installed. Please install it:\npip install openpyxl")
            return
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Bank Statement Excel File", "",
            "Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        if not file_path:
            return
        self.file_path_edit.setText(file_path)
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                QMessageBox.warning(self, "Empty File", "The Excel file appears empty.")
                return
            headers = [str(cell) if cell is not None else "" for cell in rows[0]]
            headers = [h.strip() for h in headers]
            data_rows = []
            for row in rows[1:]:
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                data_rows.append(row)
            if not data_rows:
                QMessageBox.warning(self, "No Data", "No data rows found after the header.")
                return
            ledger_cols = [i for i, h in enumerate(headers) if re.search(r"ledger\s*\d+", h, re.IGNORECASE)]
            amount_cols = [i for i, h in enumerate(headers) if re.search(r"amount\s*\d+", h, re.IGNORECASE)]
            pair_count = min(len(ledger_cols), len(amount_cols))
            if pair_count == 0:
                pair_count = 1
            self.pair_count = pair_count
            self.setup_table()
            mapping = self.auto_detect_mapping(headers)
            if not mapping or len(mapping) < 3:
                dlg = ColumnMappingDialog(self, headers, data_rows[:5])
                if dlg.exec() != QDialog.DialogCode.Accepted:
                    return
                mapping = dlg.get_mapping()
            self.mapping = mapping
            self.statement_data = self.parse_data(data_rows, headers, mapping)
            if not self.statement_data:
                retry = QMessageBox.question(
                    self,
                    "Parsing Error",
                    "Could not parse data using the selected mapping.\n\n"
                    "Would you like to try manual mapping again?",
                    QMessageBox.Yes | QMessageBox.No
                )
                if retry == QMessageBox.Yes:
                    dlg = ColumnMappingDialog(self, headers, data_rows[:5])
                    if dlg.exec() == QDialog.DialogCode.Accepted:
                        mapping = dlg.get_mapping()
                        self.mapping = mapping
                        self.statement_data = self.parse_data(data_rows, headers, mapping)
                if not self.statement_data:
                    QMessageBox.warning(
                        self,
                        "Parsing Error",
                        "Could not parse data. Please check your file format.\n"
                        "Use the 'Download Template' button to get the correct structure."
                    )
                    return
            self.populate_table()
            self.count_lbl.setText(f"{len(self.statement_data)} rows loaded")
            self.narration_filter_edit.clear()
            self.on_narration_filter_changed("")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read Excel file:\n{str(e)}")

    def auto_detect_mapping(self, headers):
        mapping = {}
        field_keywords = {
            "Date": ["transaction date", "date", "trans date", "posting date"],
            "Vch No.": ["vch no", "voucher no", "ref no", "reference", "chq no", "cheque no"],
            "Description": ["narration", "description", "particulars", "details"],
            "Debit": ["debit", "dr", "withdrawal"],
            "Credit": ["credit", "cr", "deposit"],
            "Balance": ["balance", "bal", "closing balance"]
        }
        for field, keywords in field_keywords.items():
            for idx, h in enumerate(headers):
                if any(k in h.lower() for k in keywords):
                    mapping[field] = idx
                    break
        if "Date" in mapping and "Description" in mapping and ("Debit" in mapping or "Credit" in mapping):
            return mapping
        else:
            return {}

    def parse_data(self, data_rows, headers, mapping):
        parsed = []
        date_col = mapping.get("Date")
        desc_col = mapping.get("Description")
        debit_col = mapping.get("Debit")
        credit_col = mapping.get("Credit")
        bal_col = mapping.get("Balance")
        ledger_cols = [i for i, h in enumerate(headers) if re.search(r"ledger\s*\d+", h, re.IGNORECASE)]
        amount_cols = [i for i, h in enumerate(headers) if re.search(r"amount\s*\d+", h, re.IGNORECASE)]
        min_pairs = min(len(ledger_cols), len(amount_cols))
        ledger_cols = ledger_cols[:min_pairs]
        amount_cols = amount_cols[:min_pairs]
        for row in data_rows:
            try:
                date_val = row[date_col] if date_col is not None and date_col < len(row) else None
                desc_val = row[desc_col] if desc_col is not None and desc_col < len(row) else None
                debit_val = row[debit_col] if debit_col is not None and debit_col < len(row) else None
                credit_val = row[credit_col] if credit_col is not None and credit_col < len(row) else None
                bal_val = row[bal_col] if bal_col is not None and bal_col < len(row) else None
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%d-%m-%Y")
                elif isinstance(date_val, str):
                    try:
                        dt = datetime.strptime(date_val, "%Y-%m-%d")
                        date_str = dt.strftime("%d-%m-%Y")
                    except:
                        date_str = date_val
                else:
                    try:
                        if isinstance(date_val, (int, float)):
                            dt = openpyxl.utils.datetime.from_excel(date_val)
                            date_str = dt.strftime("%d-%m-%Y")
                        else:
                            date_str = str(date_val)
                    except:
                        date_str = str(date_val)
                debit = clean_float(debit_val) if debit_val is not None else 0.0
                credit = clean_float(credit_val) if credit_val is not None else 0.0
                balance = clean_float(bal_val) if bal_val is not None else 0.0
                if debit == 0 and credit == 0:
                    continue
                ledger_entries = []
                for l_col, a_col in zip(ledger_cols, amount_cols):
                    led_name = row[l_col] if l_col < len(row) and row[l_col] else ""
                    amt_val = row[a_col] if a_col < len(row) and row[a_col] else 0.0
                    ledger_entries.append({
                        "ledger": str(led_name).strip(),
                        "amount": clean_float(amt_val)
                    })
                while len(ledger_entries) < self.pair_count:
                    ledger_entries.append({"ledger": "", "amount": 0.0})
                net = credit - debit
                vtype = "Receipt" if net > 0 else "Payment" if net < 0 else "Contra"
                parsed.append({
                    "date": date_str,
                    "description": str(desc_val).strip() if desc_val else "",
                    "debit": debit,
                    "credit": credit,
                    "balance": balance,
                    "ledger_entries": ledger_entries,
                    "vtype": vtype,
                })
            except Exception:
                continue
        return parsed

    def populate_table(self):
        self.table.blockSignals(True)
        self.table.setRowCount(len(self.statement_data) + 1)
        self.setup_table()
        self.posted_rows.clear()
        for idx, row_data in enumerate(self.statement_data, start=1):
            chk = QCheckBox()
            chk.setChecked(True)
            chk.setStyleSheet("margin-left: 10px;")
            self.table.setCellWidget(idx, 0, chk)
            sr_item = QTableWidgetItem(str(idx))
            sr_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            sr_item.setFlags(sr_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(idx, 1, sr_item)
            date_item = QTableWidgetItem(row_data["date"])
            date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(idx, 2, date_item)
            narr_item = QTableWidgetItem(row_data["description"])
            narr_item.setFlags(narr_item.flags() | Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(idx, 3, narr_item)
            dr_item = QTableWidgetItem(self.format_indian_currency(row_data["debit"], blank_if_zero=True))
            dr_item.setFlags(dr_item.flags() | Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(idx, 4, dr_item)
            cr_item = QTableWidgetItem(self.format_indian_currency(row_data["credit"], blank_if_zero=True))
            cr_item.setFlags(cr_item.flags() | Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(idx, 5, cr_item)
            vtype_item = QTableWidgetItem(row_data.get("vtype", "Payment"))
            vtype_item.setFlags(vtype_item.flags() | Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(idx, 6, vtype_item)
            ledger_entries = row_data.get("ledger_entries", [])
            while len(ledger_entries) < self.pair_count:
                ledger_entries.append({"ledger": "", "amount": 0.0})
            for pair_idx in range(self.pair_count):
                ledger_col = 7 + pair_idx * 2
                amount_col = 8 + pair_idx * 2
                entry = ledger_entries[pair_idx] if pair_idx < len(ledger_entries) else {"ledger": "", "amount": 0.0}
                led_item = QTableWidgetItem(entry["ledger"])
                led_item.setData(Qt.UserRole, entry["ledger"])
                self.table.setItem(idx, ledger_col, led_item)
                amt_item = QTableWidgetItem(self.format_indian_currency(entry["amount"], blank_if_zero=True))
                amt_item.setData(Qt.UserRole, entry["amount"])
                amt_item.setFlags(amt_item.flags() | Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(idx, amount_col, amt_item)
            self.update_row_balance(idx)
        self.table.blockSignals(False)

    def update_row_balance(self, row):
        if self._balancing:
            return
        debit = clean_float(self.get_cell_text(row, 4))
        credit = clean_float(self.get_cell_text(row, 5))
        net = abs(credit - debit)
        other_amt = 0.0
        for a_col in self.amount_col_indices:
            if a_col == 8:
                continue
            amt = clean_float(self.get_cell_text(row, a_col))
            other_amt += amt
        bal_fig = net - other_amt
        amt1_item = self.table.item(row, 8)
        if amt1_item:
            current_val = clean_float(amt1_item.text())
            if abs(current_val - bal_fig) > 0.01:
                self._balancing = True
                amt1_item.setText(self.format_indian_currency(bal_fig, blank_if_zero=False))
                self._balancing = False
        total_amt = 0.0
        for a_col in self.amount_col_indices:
            amt = clean_float(self.get_cell_text(row, a_col))
            total_amt += amt
        is_bal = abs(total_amt - net) < 0.01
        if row in getattr(self, 'posted_rows', set()):
            bg_color = QColor("#d4edda")
        elif not is_bal:
            bg_color = QColor("#ffe6e6")
        else:
            bg_color = QColor("#ffffff" if row % 2 == 0 else "#faf9f4")
        for c in range(self.table.columnCount()):
            it = self.table.item(row, c)
            if not it:
                it = QTableWidgetItem("")
                self.table.setItem(row, c, it)
            it.setBackground(QBrush(bg_color))
        chk = self.table.cellWidget(row, 0)
        if isinstance(chk, QCheckBox):
            chk.setStyleSheet(f"margin-left: 10px; background-color: {bg_color.name()};")

    def on_cell_changed(self, row, col):
        if row <= 0:
            return
        if hasattr(self, 'posted_rows') and row in self.posted_rows:
            self.posted_rows.discard(row)
            self.update_row_balance(row)
        if col in (5, 6) or col in self.amount_col_indices:
            self.update_row_balance(row)

    def format_indian_currency(self, number, blank_if_zero=False):
        if number is None or (isinstance(number, float) and number != number) or number == "":
            return ""
        try:
            num = float(number)
            if blank_if_zero and num == 0:
                return ""
            is_neg = num < 0
            s = f"{abs(num):.2f}"
            int_part, dec_part = s.split('.')
            if len(int_part) <= 3:
                result = int_part
            else:
                last_three = int_part[-3:]
                remaining = int_part[:-3]
                remaining = re.sub(r'(\d+?)(?=(\d\d)+$)', r'\1,', remaining)
                result = remaining + ',' + last_three
            final = result + '.' + dec_part
            return '-' + final if is_neg else final
        except:
            return str(number)

    def save_to_excel(self):
        if self.table.rowCount() <= 1:
            QMessageBox.information(self, "Save", "No data to save.")
            return
        headers = []
        for c in range(self.table.columnCount()):
            headers.append(self.table.horizontalHeaderItem(c).text())
        data = []
        for r in range(1, self.table.rowCount()):
            row_data = []
            for c in range(self.table.columnCount()):
                it = self.table.item(r, c)
                if it:
                    row_data.append(it.text())
                else:
                    w = self.table.cellWidget(r, c)
                    if isinstance(w, QComboBox):
                        row_data.append(w.currentText())
                    else:
                        row_data.append("")
            data.append(row_data)
        wb = Workbook()
        ws = wb.active
        ws.title = "Bank Import Data"
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        for row_idx, row in enumerate(data, 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[col_letter].width = adjusted_width
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Bank Import Data",
            "Bank_Import_Backup.xlsx",
            "Excel Files (*.xlsx)"
        )
        if file_path:
            try:
                wb.save(file_path)
                QMessageBox.information(self, "Success", f"Data saved to:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Could not save file:\n{str(e)}")

    def clear_data(self):
        self.statement_data = []
        self.table.setRowCount(1)
        self.count_lbl.setText("0 rows loaded")
        self.file_path_edit.clear()
        self.narration_filter_edit.clear()
        self.on_narration_filter_changed("")

    def select_all(self):
        for r in range(1, self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if isinstance(w, QCheckBox):
                w.setChecked(True)

    def deselect_all(self):
        for r in range(1, self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if isinstance(w, QCheckBox):
                w.setChecked(False)

    def get_cell_text(self, row, col):
        it = self.table.item(row, col)
        return it.text().strip() if it else ""

    def on_header_clicked(self, col_idx):
        if col_idx > 0:
            self.open_excel_column_filter(col_idx)

    def on_header_context_menu(self, pos):
        col_idx = self.table.horizontalHeader().logicalIndexAt(pos)
        if col_idx < 0:
            return
        col_name = self.table.horizontalHeaderItem(col_idx).text()
        menu = QMenu(self)
        menu.setStyleSheet("font-family: 'Lexend', 'Segoe UI'; font-size: 11px;")
        act_filter = menu.addAction(f"🔍 Filter & Sort '{col_name}'...")
        act_filter.triggered.connect(lambda: self.open_excel_column_filter(col_idx))
        menu.addSeparator()
        act_hide = menu.addAction(f"👁 Hide Column '{col_name}'")
        act_hide.triggered.connect(lambda: self.hide_column(col_idx))
        act_unhide_all = menu.addAction("👁 Unhide All Columns")
        act_unhide_all.triggered.connect(self.unhide_all_columns)
        menu.addSeparator()
        act_manage = menu.addAction("⚙ Column Visibility Manager...")
        act_manage.triggered.connect(self.open_column_manager)
        if col_idx in self.column_filters:
            act_clear_filter = menu.addAction(f"Clear Filter on '{col_name}'")
            act_clear_filter.triggered.connect(lambda: self.clear_column_filter(col_idx))
        menu.exec(self.table.horizontalHeader().mapToGlobal(pos))

    def hide_column(self, col_idx):
        self.hidden_columns.add(col_idx)
        self.table.setColumnHidden(col_idx, True)

    def unhide_all_columns(self):
        self.hidden_columns.clear()
        for c in range(self.table.columnCount()):
            self.table.setColumnHidden(c, False)

    def open_column_manager(self):
        column_headers = [(self.table.horizontalHeaderItem(i).text(), self.table.columnWidth(i)) for i in range(self.table.columnCount())]
        dlg = ColumnVisibilityDialog(self, column_headers, self.hidden_columns)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.hidden_columns = dlg.hidden_columns
            for c in range(self.table.columnCount()):
                self.table.setColumnHidden(c, c in self.hidden_columns)

    def open_excel_column_filter(self, col_idx):
        col_title = self.table.horizontalHeaderItem(col_idx).text()
        unique_vals = set()
        is_numeric = col_idx in self.amount_col_indices or col_idx in (4,5)
        for r in range(1, self.table.rowCount()):
            it = self.table.item(r, col_idx)
            val = it.text() if it else ""
            unique_vals.add(val)
        sorted_vals = sorted(list(unique_vals), key=str.lower)
        curr_selected = self.column_filters.get(col_idx, None)
        dlg = ExcelColumnFilterDialog(self, col_title, sorted_vals, curr_selected, is_numeric)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            if dlg.sort_order:
                self.sort_data_rows(col_idx, dlg.sort_order)
            if dlg.selected_values is None:
                self.clear_column_filter(col_idx)
            else:
                if is_numeric and dlg.min_val is not None and dlg.max_val is not None:
                    self.column_filters[col_idx] = {"type": "range", "min": dlg.min_val, "max": dlg.max_val}
                else:
                    self.column_filters[col_idx] = {"type": "values", "values": dlg.selected_values}
                self.update_header_labels()
                self.apply_column_filters()

    def sort_data_rows(self, col_idx, order):
        if self.table.rowCount() <= 1:
            return
        rows_data = []
        for r in range(1, self.table.rowCount()):
            chk_widget = self.table.cellWidget(r, 0)
            is_checked = chk_widget.isChecked() if isinstance(chk_widget, QCheckBox) else True
            row_items = []
            for c in range(self.table.columnCount()):
                it = self.table.item(r, c)
                if it:
                    row_items.append(it.text())
                else:
                    w = self.table.cellWidget(r, c)
                    if isinstance(w, QComboBox):
                        row_items.append(w.currentText())
                    else:
                        row_items.append("")
            rows_data.append({"checked": is_checked, "items": row_items, "row_index": r})
        if not rows_data:
            return
        def get_sort_key(row_dict):
            val = row_dict["items"][col_idx] if col_idx < len(row_dict["items"]) else ""
            if col_idx in self.amount_col_indices or col_idx in (4,5):
                cleaned = re.sub(r"[^d.-]", "", val)
                if cleaned:
                    try:
                        return (0, float(cleaned))
                    except ValueError:
                        pass
            if col_idx == 2 and val:
                parts = val.split("-")
                if len(parts) == 3:
                    return (0, f"{parts[2]}{parts[1]}{parts[0]}")
            return (1, val.lower())
        reverse = (order == "DESC")
        rows_data.sort(key=get_sort_key, reverse=reverse)
        self.table.blockSignals(True)
        for new_row, r_data in enumerate(rows_data, start=1):
            for c, text_val in enumerate(r_data["items"]):
                it = self.table.item(new_row, c)
                if it:
                    it.setText(text_val)
                else:
                    it = QTableWidgetItem(text_val)
                    self.table.setItem(new_row, c, it)
            chk = self.table.cellWidget(new_row, 0)
            if isinstance(chk, QCheckBox):
                chk.blockSignals(True)
                chk.setChecked(r_data["checked"])
                chk.blockSignals(False)
        self.table.blockSignals(False)
        self.apply_column_filters()
        self.log(f"Sorted by column {col_idx} ({order})")

    def clear_column_filter(self, col_idx):
        if col_idx in self.column_filters:
            del self.column_filters[col_idx]
            self.update_header_labels()
            self.apply_column_filters()

    def update_header_labels(self):
        for col_idx in range(self.table.columnCount()):
            item = self.table.horizontalHeaderItem(col_idx)
            if item:
                orig_text = item.text().replace(" [▼]", "")
                if col_idx in self.column_filters:
                    item.setText(f"{orig_text} [▼]")
                else:
                    item.setText(orig_text)

    def apply_column_filters(self):
        for r in range(1, self.table.rowCount()):
            row_visible = True
            for col_idx, filter_info in self.column_filters.items():
                it = self.table.item(r, col_idx)
                val = it.text() if it else ""
                if filter_info.get("type") == "range":
                    try:
                        num_val = float(val.replace(',', '')) if val else 0.0
                        if not (filter_info["min"] <= num_val <= filter_info["max"]):
                            row_visible = False
                            break
                    except:
                        row_visible = False
                        break
                else:
                    if val not in filter_info.get("values", set()):
                        row_visible = False
                        break
            self.table.setRowHidden(r, not row_visible)

    def get_validation_errors(self):
        """Return list of error strings for all rows."""
        errors = []
        for r in range(1, self.table.rowCount()):
            chk = self.table.cellWidget(r, 0)
            if not isinstance(chk, QCheckBox) or not chk.isChecked():
                continue
                
            row_errors = []
            
            date_str = self.get_cell_text(r, 2)
            if date_str and hasattr(self.parent_app, 'period_from') and self.parent_app.period_from:
                d_tally = format_date_to_tally(date_str)
                if not (self.parent_app.period_from <= d_tally <= self.parent_app.period_to):
                    row_errors.append(f"Row {r}: Date {date_str} is outside the selected period ({self.parent_app.period_from} to {self.parent_app.period_to})")

            debit = clean_float(self.get_cell_text(r, 4))
            credit = clean_float(self.get_cell_text(r, 5))
            if debit == 0 and credit == 0:
                row_errors.append(f"Row {r}: Credit or Debit amount must be greater than zero.")
            net = credit - debit
            total_amt = 0.0
            for a_col in self.amount_col_indices:
                amt = clean_float(self.get_cell_text(r, a_col))
                total_amt += amt
            if abs(total_amt - abs(net)) > 0.01:
                row_errors.append(f"Row {r}: Sum of amounts ({total_amt:.2f}) does not match net amount ({abs(net):.2f})")

            # Check: if amount > 0, ledger must be present
            for l_col, a_col in zip(self.ledger_col_indices, self.amount_col_indices):
                amt = clean_float(self.get_cell_text(r, a_col))
                if amt > 0:
                    led_name = self.get_cell_text(r, l_col)
                    if not led_name:
                        row_errors.append(f"Row {r}: Amount {amt:.2f} present but no ledger selected in column {l_col+1}")

            # Check ledger exists in Tally masters
            for l_col in self.ledger_col_indices:
                led_name = self.get_cell_text(r, l_col)
                if led_name:
                    if not any(led_name.lower() == l.lower() for l in self.parent_app.ledgers):
                        row_errors.append(f"Row {r}: Ledger '{led_name}' not found in Tally masters")

            if row_errors:
                errors.extend(row_errors)
        return errors

    def validate_data(self):
        errors = self.get_validation_errors()
        if errors:
            QMessageBox.critical(self, "Validation Failed", "\n".join(errors))
        else:
            QMessageBox.information(self, "Validation Passed", "All rows are balanced and all ledgers are valid.")

    def get_selected_tasks(self):
        tasks = []
        for r in range(1, self.table.rowCount()):
            chk = self.table.cellWidget(r, 0)
            if not isinstance(chk, QCheckBox) or not chk.isChecked():
                continue
            date_str = self.get_cell_text(r, 2)
            vch_no = self.get_cell_text(r, 3)
            narration = self.get_cell_text(r, 4)
            debit = clean_float(self.get_cell_text(r, 5))
            credit = clean_float(self.get_cell_text(r, 6))
            vch_type = self.get_cell_text(r, 7)
            if debit > 0 and vch_type not in ["Payment", "Contra"]:
                vch_type = "Payment"
            elif credit > 0 and vch_type not in ["Receipt", "Contra"]:
                vch_type = "Receipt"
            elif debit == 0 and credit == 0:
                continue
            if vch_type == "Payment":
                bank_amount = debit
            elif vch_type == "Receipt":
                bank_amount = credit
            else:
                bank_amount = debit or credit
            entries = []
            for l_col, a_col in zip(self.ledger_col_indices, self.amount_col_indices):
                led_name = self.get_cell_text(r, l_col)
                amt = clean_float(self.get_cell_text(r, a_col))
                if led_name and amt > 0:
                    if vch_type in ["Payment", "Contra"]:
                        if led_name == self.bank_ledger:
                            continue
                        else:
                            entries.append({"name": led_name, "deemed_positive": "Yes", "amount": f"-{amt:.2f}"})
                    else:
                        if led_name == self.bank_ledger:
                            continue
                        else:
                            entries.append({"name": led_name, "deemed_positive": "No", "amount": f"{amt:.2f}"})
            if vch_type in ["Payment", "Contra"]:
                entries.append({"name": self.bank_ledger, "deemed_positive": "No", "amount": f"{bank_amount:.2f}"})
            else:
                entries.append({"name": self.bank_ledger, "deemed_positive": "Yes", "amount": f"-{bank_amount:.2f}"})
            tasks.append({
                "vch_type": vch_type,
                "date": date_str,
                "vch_no": vch_no,
                "narration": narration,
                "ledger_entries": entries,
                "party_ledger": "",
                "row": r
            })
        return tasks

    def on_export_clicked(self):
        errors = self.get_validation_errors()
        if errors:
            QMessageBox.critical(self, "Validation Failed", "Please fix the following errors before exporting:\n\n" + "\n".join(errors))
            return
        if not self.parent_app.active_company:
            QMessageBox.warning(self, "Missing Company", "No company loaded. Please connect to Tally first.")
            return
        if not self.bank_ledger:
            QMessageBox.warning(self, "Missing Bank Ledger", "Please select a Bank Ledger from the dropdown.")
            return
            
        tasks = self.get_selected_tasks()
        if not tasks:
            QMessageBox.information(self, "No Tasks", "No valid rows selected to export.")
            return
            
        file_path, _ = QFileDialog.getSaveFileName(self, "Save XML", "", "XML Files (*.xml)")
        if not file_path:
            return

        xml_content = "<ENVELOPE>\n<BODY>\n<IMPORTDATA>\n<REQUESTDATA>\n<TALLYMESSAGE xmlns:UDF=\"TallyUDF\">\n"
        for t in tasks:
            req = build_create_voucher_request(self.parent_app.active_company, t['vch_type'], format_date_to_tally(t['date']), t['narration'], t['ledger_entries'], t.get('party_ledger', ''))
            m = re.search(r'(<VOUCHER .*?</VOUCHER>)', req, flags=re.DOTALL)
            if m:
                xml_content += m.group(1) + "\n"
        xml_content += "</TALLYMESSAGE>\n</REQUESTDATA>\n</IMPORTDATA>\n</BODY>\n</ENVELOPE>"
        
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(xml_content)
        self.log(f"Exported {len(tasks)} vouchers to {file_path}", "OK")
        QMessageBox.information(self, "Export Complete", f"Successfully exported {len(tasks)} vouchers to XML.")

    def post_selected(self):
        # First, run validation and prevent posting if errors exist
        errors = self.get_validation_errors()
        if errors:
            QMessageBox.critical(self, "Validation Failed", "Please fix the following errors before posting:\n\n" + "\n".join(errors))
            return

        if not self.parent_app.active_company:
            QMessageBox.warning(self, "Missing Company", "No company loaded. Please connect to Tally first.")
            return
        if not self.bank_ledger:
            QMessageBox.warning(self, "Missing Bank Ledger", "Please select a Bank Ledger from the dropdown.")
            return

        tasks = self.get_selected_tasks()
        if not tasks:
            QMessageBox.information(self, "No Tasks", "No valid rows selected.")
            return
        reply = QMessageBox.question(
            self,
            "Confirm Post",
            f"Ready to create {len(tasks)} voucher(s) in '{self.parent_app.active_company}'.\nProceed?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        self.btn_post.setEnabled(False)
        self.post_worker = TallyCreateWorker(
            self.parent_app.url_edit.text().strip(),
            self.parent_app.active_company,
            tasks
        )
        self.post_worker.progress.connect(lambda cur, tot, msg: None)
        self.post_worker.finished.connect(self.on_post_finished)
        self.post_worker.log_msg.connect(self.on_post_log)
        self.post_worker.start()

    def on_post_log(self, msg, level):
        self.log(msg, level)
        if not hasattr(self, '_post_logs'):
            self._post_logs = []
        self._post_logs.append(f"[{level}] {msg}")

    def on_post_finished(self, success_count, total):
        self.btn_post.setEnabled(True)
        for task in getattr(self.post_worker, 'tasks', []):
            if task.get("success"):
                row = task.get('row')
                if row:
                    self.posted_rows.add(row)
                    for c in range(self.table.columnCount()):
                        it = self.table.item(row, c)
                        if not it:
                            it = QTableWidgetItem("")
                            self.table.setItem(row, c, it)
                        it.setBackground(QBrush(QColor("#d4edda")))
                    chk = self.table.cellWidget(row, 0)
                    if isinstance(chk, QCheckBox):
                        chk.setStyleSheet("margin-left: 10px; background-color: #d4edda;")
        self.log(f"Batch posting complete: {success_count}/{total} vouchers created successfully.", "OK")
        self._post_logs = []

# ============================================================================
#  8. DELEGATE FOR BANK IMPORT – with double‑dropdown fix and calendar styles
# ============================================================================
class BankImportDelegate(QStyledItemDelegate):
    def __init__(self, parent=None, ledgers=None, voucher_types=None):
        super().__init__(parent)
        self.ledgers = ledgers or []
        self.voucher_types = voucher_types or []

    def set_ledgers(self, ledgers):
        self.ledgers = ledgers

    def set_voucher_types(self, voucher_types):
        self.voucher_types = voucher_types

    def paint(self, painter, option, index):
        super().paint(painter, option, index)
        col = index.column()
        row = index.row()
        if row > 0 and (col == 7 or (col >= 8 and (col - 8) % 2 == 0)):
            painter.save()
            painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
            arrow_w = 8
            arrow_h = 5
            x = option.rect.right() - 14
            y = option.rect.top() + (option.rect.height() - arrow_h) // 2
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QBrush(QColor("#6b8079")))
            points = [
                QPoint(x, y),
                QPoint(x + arrow_w, y),
                QPoint(x + arrow_w // 2, y + arrow_h)
            ]
            painter.drawPolygon(QPolygon(points))
            painter.restore()

    def createEditor(self, parent, option, index):
        col = index.column()
        row = index.row()
        if row == 0:
            return None
        if col == 7:  # VType
            combo = SearchableComboBox(parent)
            combo.setEditable(True)
            combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
            combo.setStyleSheet("""
                QComboBox {
                    border: 2px solid #107c41;
                    background-color: #ffffff;
                    font-family: 'Lexend', 'Segoe UI', sans-serif;
                    font-size: 11px;
                    padding: 1px 3px;
                }
                QComboBox::drop-down {
                    subcontrol-origin: padding;
                    subcontrol-position: top right;
                    width: 18px;
                    border-left: 1px solid #c4bfae;
                    background-color: #f3f1ea;
                }
                QComboBox::down-arrow {
                    image: none;
                    width: 0;
                    height: 0;
                    border-left: 4px solid transparent;
                    border-right: 4px solid transparent;
                    border-top: 5px solid #0f3d33;
                }
                QComboBox QAbstractItemView {
                    border: 1px solid #107c41;
                    background-color: #ffffff;
                    color: #10221c;
                    selection-background-color: #107c41;
                    selection-color: #ffffff;
                    font-family: 'Lexend', 'Segoe UI', sans-serif;
                    font-size: 11px;
                }
            """)
            combo.addItem("")
            for vt in self.voucher_types:
                combo.addItem(vt)
            completer = QCompleter(self.voucher_types, combo)
            completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
            completer.setFilterMode(Qt.MatchFlag.MatchContains)
            combo.setCompleter(completer)
            return combo
        elif col >= 8 and (col - 8) % 2 == 0:  # ledger columns
            combo = SearchableComboBox(parent)
            combo.setEditable(True)
            combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
            combo.setStyleSheet("""
                QComboBox {
                    border: 2px solid #107c41;
                    background-color: #ffffff;
                    font-family: 'Lexend', 'Segoe UI', sans-serif;
                    font-size: 11px;
                    padding: 1px 3px;
                }
                QComboBox::drop-down {
                    subcontrol-origin: padding;
                    subcontrol-position: top right;
                    width: 18px;
                    border-left: 1px solid #c4bfae;
                    background-color: #f3f1ea;
                }
                QComboBox::down-arrow {
                    image: none;
                    width: 0;
                    height: 0;
                    border-left: 4px solid transparent;
                    border-right: 4px solid transparent;
                    border-top: 5px solid #0f3d33;
                }
                QComboBox QAbstractItemView {
                    border: 1px solid #107c41;
                    background-color: #ffffff;
                    color: #10221c;
                    selection-background-color: #107c41;
                    selection-color: #ffffff;
                    font-family: 'Lexend', 'Segoe UI', sans-serif;
                    font-size: 11px;
                }
            """)
            combo.addItem("")
            for led in self.ledgers:
                combo.addItem(led)
            completer = QCompleter(self.ledgers, combo)
            completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
            completer.setFilterMode(Qt.MatchFlag.MatchContains)
            combo.setCompleter(completer)
            return combo
        elif col >= 9 and (col - 9) % 2 == 0:  # amount columns
            line_edit = QLineEdit(parent)
            line_edit.setStyleSheet("""
                QLineEdit {
                    border: 2px solid #107c41;
                    background-color: #ffffff;
                    font-family: 'Lexend', 'Segoe UI', sans-serif;
                    font-size: 11px;
                    padding: 1px 4px;
                }
            """)
            return line_edit
        elif col == 2:  # Date
            date_edit = QDateEdit(parent)
            date_edit.setCalendarPopup(True)
            date_edit.setDisplayFormat("dd-MM-yyyy")
            date_edit.setStyleSheet("""
                QDateEdit {
                    border: 2px solid #107c41;
                    background-color: #ffffff;
                    color: #10221c;
                    font-family: 'Lexend', 'Segoe UI', sans-serif;
                    font-size: 11px;
                    padding: 1px 3px;
                }
                QDateEdit QCalendarWidget {
                    background-color: white;
                    color: black;
                }
            """)
            cal = date_edit.calendarWidget()
            if cal:
                cal.setStyleSheet("""
                    QCalendarWidget {
                        background-color: white;
                        color: black;
                    }
                    QCalendarWidget QAbstractItemView {
                        background-color: white;
                        color: black;
                    }
                    QCalendarWidget QAbstractItemView:selected {
                        background-color: #107c41;
                        color: white;
                    }
                    QCalendarWidget QToolButton {
                        color: black;
                        background-color: white;
                    }
                    QCalendarWidget QMenu {
                        color: black;
                        background-color: white;
                    }
                    QCalendarWidget QSpinBox {
                        color: black;
                        background-color: white;
                    }
                """)
            return date_edit
        elif col == 3:  # Narration
            line_edit = QLineEdit(parent)
            line_edit.setStyleSheet("""
                QLineEdit {
                    border: 2px solid #107c41;
                    background-color: #ffffff;
                    font-family: 'Lexend', 'Segoe UI', sans-serif;
                    font-size: 11px;
                    padding: 1px 4px;
                }
            """)
            return line_edit
        return super().createEditor(parent, option, index)

    def setEditorData(self, editor, index):
        val = index.model().data(index, Qt.ItemDataRole.EditRole)
        if isinstance(editor, QDateEdit):
            if val is not None:
                d = QDate.fromString(str(val), "dd-MM-yyyy")
                if not d.isValid():
                    d = QDate.fromString(str(val), "yyyyMMdd")
                if d.isValid():
                    editor.setDate(d)
        elif isinstance(editor, QComboBox):
            if val is not None:
                editor.setCurrentText(str(val))
                if editor.lineEdit() and editor.isEditable():
                    editor.lineEdit().selectAll()
        elif isinstance(editor, QLineEdit):
            if val is not None:
                editor.setText(str(val))
                editor.deselect()
                editor.setCursorPosition(len(editor.text()))
        else:
            super().setEditorData(editor, index)

    def setModelData(self, editor, model, index):
        if isinstance(editor, QDateEdit):
            model.setData(index, editor.date().toString("dd-MM-yyyy"), Qt.ItemDataRole.EditRole)
        elif isinstance(editor, QComboBox):
            txt = editor.currentText().strip()
            col = index.column()
            if col >= 8 and (col - 8) % 2 == 0:
                if txt:
                    matched = next((l for l in self.ledgers if l.lower() == txt.lower()), None)
                    if matched:
                        model.setData(index, matched, Qt.ItemDataRole.EditRole)
                    else:
                        return
                else:
                    model.setData(index, "", Qt.ItemDataRole.EditRole)
            else:
                if txt in self.voucher_types:
                    model.setData(index, txt, Qt.ItemDataRole.EditRole)
                else:
                    model.setData(index, "", Qt.ItemDataRole.EditRole)
        elif isinstance(editor, QLineEdit):
            txt = editor.text().strip()
            col = index.column()
            if col >= 9 and (col - 9) % 2 == 0:
                try:
                    cleaned = ''.join(c for c in txt if c.isdigit() or c == '.' or c == '-')
                    if cleaned:
                        val = float(cleaned)
                        formatted = format_indian_currency(abs(val), blank_if_zero=True)
                        model.setData(index, formatted, Qt.ItemDataRole.EditRole)
                    else:
                        model.setData(index, "", Qt.ItemDataRole.EditRole)
                except:
                    model.setData(index, txt, Qt.ItemDataRole.EditRole)
            else:
                model.setData(index, txt, Qt.ItemDataRole.EditRole)
        else:
            super().setModelData(editor, model, index)

# ============================================================================
#  9. MAIN WINDOW WITH TABS
# ============================================================================
class TallyEditorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Tally Voucher Editor & Bank Import (Tabbed)")
        self.setMinimumSize(800, 600)
        self.resize(1400, 820)
        self.showMaximized()

        self.active_company = ""
        self.ledgers = []
        self.voucher_types = []

        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #d8d2c1;
                background: #ffffff;
            }
            QTabBar::tab {
                font-family: 'Lexend';
                font-size: 12px;
                padding: 6px 12px;
                background: #f3f1ea;
                border: 1px solid #c4bfae;
                border-bottom: none;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background: #ffffff;
                border-bottom: 2px solid #107c41;
            }
            QTabBar::tab:hover {
                background: #e5e0d3;
            }
        """)
        self.setCentralWidget(self.tab_widget)

        self.editor_widget = VoucherEditorWidget(self)
        self.tab_widget.addTab(self.editor_widget, "📝 Voucher Editor")

        self.url_edit = self.editor_widget.url_edit

        self.bank_widget = BankImportWidget(self)
        self.tab_widget.addTab(self.bank_widget, "🏦 Bank Import")
        self.bank_widget.setEnabled(False)

        self.btn_help = QPushButton("ℹ Help / Instructions")
        self.btn_help.setStyleSheet("background-color: #0f3d33; color: #ffffff; padding: 4px 12px; font-weight: bold; border-radius: 4px; border: none;")
        self.btn_help.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_help.clicked.connect(self.show_help)
        self.tab_widget.setCornerWidget(self.btn_help, Qt.Corner.TopRightCorner)

        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

        QTimer.singleShot(500, self.show_startup_message)
        QTimer.singleShot(2000, self.check_updates)

    def check_updates(self):
        self.update_checker = UpdateCheckerWorker()
        self.update_checker.update_available.connect(self.on_update_available)
        self.update_checker.start()

    def on_update_available(self, version, download_url):
        reply = QMessageBox.question(self, "Update Available",
                                     f"A new version ({version}) of Tally Integration Pro is available!\n\nWould you like to download and install it now?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            from PySide6.QtWidgets import QProgressDialog
            self.progress_dialog = QProgressDialog("Downloading update...", "Cancel", 0, 100, self)
            self.progress_dialog.setWindowTitle("Updating")
            self.progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
            self.progress_dialog.show()

            self.downloader = UpdateDownloaderWorker(download_url)
            self.downloader.progress.connect(self.progress_dialog.setValue)
            self.downloader.finished_download.connect(self.apply_update)
            self.downloader.error.connect(lambda e: QMessageBox.critical(self, "Update Failed", f"Failed to download: {e}"))
            self.downloader.start()

    def apply_update(self, temp_path):
        self.progress_dialog.setValue(100)
        import sys, os, subprocess
        current_exe = sys.executable
        if not current_exe.endswith(".exe") or "python" in current_exe.lower():
            QMessageBox.information(self, "Development Mode", "Update downloaded successfully, but cannot be applied automatically while running from source code.")
            return

        bat_path = os.path.join(os.environ.get('TEMP', ''), "update_tally.bat")
        bat_content = f"""@echo off
timeout /t 2 /nobreak > NUL
move /y "{temp_path}" "{current_exe}"
start "" "{current_exe}"
del "%~f0"
"""
        with open(bat_path, "w") as f:
            f.write(bat_content)
        
        subprocess.Popen([bat_path], shell=True, creationflags=subprocess.CREATE_NO_WINDOW)
        sys.exit()

    def show_startup_message(self):
        msg = QMessageBox(self)
        msg.setWindowTitle("Welcome to Bank Entry Tool!")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setText("<b>Welcome to Bank Entry Tool!</b><br><br>"
                    "Please follow these instructions for a safe experience:<br><br>"
                    "1. First Keep the Tally Backup in a separate folder from the original folder for Safety purposes.<br><br>"
                    "2. Tally Should Run in <b>'Serve as Both'</b> mode and the Port Code must be <b>9000</b>.<br><br>"
                    "<i>To read detailed instructions, click the 'Help / Instructions' button in the top right corner.</i>")
        msg.exec()

    def show_help(self):
        msg = QMessageBox(self)
        msg.setWindowTitle("Detailed Instructions")
        msg.setIcon(QMessageBox.Icon.Information)
        instructions = """
<h3>How to use Tally Integration Pro</h3>

<b>1. Setup & Connection</b>
- Open Tally Prime. Ensure it is running in <b>'Serve as Both'</b> mode and the port is set to <b>9000</b>.
- Keep a <b>backup</b> of your Tally data in a separate folder for safety.
- Click 'Connect' in the tool to load your Ledgers and Voucher Types.

<b>2. Bank Import</b>
- Click 'Download Template' to get the expected Excel structure.
- Fill the template with your bank statement data. You can add 'Counter Ledgers' horizontally if a transaction is split across multiple ledgers.
- Ensure all mandatory columns (*) are filled.
- Load the Excel file. Map the columns if necessary.
- Select the Bank Ledger, validate the entries, and click 'Post Selected to Tally'.

<b>3. Voucher Editor</b>
- Fetches Daybook transactions from Tally within a date range.
- You can filter, validate, edit, and post the modified vouchers back to Tally.

<hr>
<b>Privacy policy & Copyright</b>
<br>Copyright © 2026 Pratik Ashtagi. All rights reserved.
<br>Property of Pratik Ashtagi. Created and owned by Pratik Ashtagi.
<br><br><b>Support & Feedback</b>
<br>Any feedback or Suggestion pls Mail to: ashatgipratik6@gmail.com
<br>Whatsapp Mob: 9999999999
"""
        msg.setText(instructions)
        msg.exec()


    def enable_bank_import(self, enabled):
        self.bank_widget.setEnabled(enabled)
        if enabled:
            self.bank_widget.set_ledgers(self.ledgers)
            self.bank_widget.set_voucher_types(self.voucher_types)

# ============================================================================
#  10. APPLICATION ENTRY POINT
# ============================================================================
def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    font = QFont("Lexend", 9)
    if not font.exactMatch():
        font = QFont("Segoe UI", 9)
    app.setFont(font)

    window = TallyEditorApp()
    window.showMaximized()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()